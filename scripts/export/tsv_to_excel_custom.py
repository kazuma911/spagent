"""Populate a custom Excel template from TSV.

`data/excel-template-mapping.json` のセル指定に従い、メニュー TSV のヘッダ、
本文、pace table をコーチ提供テンプレートへ書き込む。
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path
from typing import Any


def repo_root() -> Path:
    """Return the repository root.

    スクリプトの場所からリポジトリルートを推定する。
    """
    return Path(__file__).resolve().parents[2]


def import_openpyxl() -> Any:
    """Import openpyxl with a helpful error.

    依存がない場合はインストール方法を含む例外にする。
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required. Install with: python -m pip install openpyxl") from exc
    return openpyxl


def normalize_key(value: str) -> str:
    """Normalize metadata keys.

    TSV コメントや列名を小文字キーにする。
    """
    return re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")


def parse_tsv(path: Path) -> tuple[dict[str, str], list[dict[str, str]], list[dict[str, str]]]:
    """Parse menu and pace table sections from TSV.

    主メニューと `# _pace_table` を別々の辞書リストとして返す。
    """
    meta: dict[str, str] = {}
    menu_lines: list[str] = []
    pace_lines: list[str] = []
    section = "menu"
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        for raw_line in file:
            line = raw_line.rstrip("\n")
            stripped = line.strip()
            if not stripped:
                continue
            if stripped.startswith("#"):
                marker = stripped.lower()
                if marker.startswith("# _pace_table"):
                    section = "pace"
                    continue
                if marker.startswith("# _"):
                    section = "other"
                    continue
                text = stripped.lstrip("#").strip()
                if ":" in text:
                    key, value = text.split(":", 1)
                    meta[normalize_key(key)] = value.strip()
                continue
            if section == "pace":
                pace_lines.append(line)
            elif section == "menu":
                menu_lines.append(line)

    def read(lines: list[str]) -> list[dict[str, str]]:
        if not lines:
            return []
        return [
            {normalize_key(key or ""): (value or "").strip() for key, value in row.items()}
            for row in csv.DictReader(lines, delimiter="\t")
        ]

    return meta, read(menu_lines), read(pace_lines)


def load_mapping(path: Path) -> dict[str, Any]:
    """Load Excel template mapping JSON.

    セル位置と列マッピングを含む JSON オブジェクトを読み込む。
    """
    with path.open("r", encoding="utf-8") as file:
        data = json.load(file)
    if not isinstance(data, dict):
        raise ValueError(f"{path} must contain a JSON object")
    return data


def write_row(ws: Any, row_index: int, columns: dict[str, str], values: dict[str, str]) -> None:
    """Write one TSV row to mapped Excel columns.

    `columns` は TSV キーから Excel 列文字への対応。
    """
    for key, column in columns.items():
        ws[f"{column}{row_index}"] = values.get(normalize_key(key), "")


def convert_to_excel(path: Path, template: Path | None, mapping_path: Path, out_path: Path | None, *, append_to: Path | None = None, new_sheet_name: str | None = None) -> Path:
    """Populate a workbook from TSV using a mapping.

    テンプレートが省略された場合は新規ワークブックを使う。
    `append_to` を指定すると既存ブックを開き、`new_sheet_name` の新シートに書き込む
    （SPA.xlsx のように 1 ブック複数シートの履歴集に追記するユースケース）。
    """
    openpyxl = import_openpyxl()
    mapping = load_mapping(mapping_path)
    meta, menu_rows, pace_rows = parse_tsv(path)
    if append_to:
        workbook = openpyxl.load_workbook(append_to)
        sheet_name = new_sheet_name or meta.get("date") or path.stem
        while sheet_name in workbook.sheetnames:
            sheet_name = f"{sheet_name}-2"
        ws = workbook.create_sheet(title=sheet_name)
    elif template:
        workbook = openpyxl.load_workbook(template)
        ws = workbook[mapping.get("sheet_name")] if mapping.get("sheet_name") else workbook.active
    else:
        workbook = openpyxl.Workbook()
        ws = workbook[mapping.get("sheet_name")] if mapping.get("sheet_name") else workbook.active
        if new_sheet_name:
            ws.title = new_sheet_name

    team_name_cell = mapping.get("team_name")
    if team_name_cell:
        ws[team_name_cell] = meta.get("team_name") or meta.get("team") or ""
    for key in ("title", "date", "facility", "theme", "equipment"):
        cell = mapping.get(key)
        if cell:
            ws[cell] = meta.get(key, "")
    for cell_addr, header_text in (mapping.get("headers") or {}).items():
        ws[cell_addr] = header_text

    body_start = int(mapping.get("body_start_row", 2))
    columns = mapping.get("columns", {})
    if not isinstance(columns, dict) or not columns:
        raise ValueError("mapping must include a non-empty 'columns' object")
    for offset, row in enumerate(menu_rows):
        write_row(ws, body_start + offset, columns, row)

    pace_start = mapping.get("pace_table_start_row")
    pace_columns = mapping.get("pace_columns", {})
    if pace_start and isinstance(pace_columns, dict):
        for offset, row in enumerate(pace_rows):
            write_row(ws, int(pace_start) + offset, pace_columns, row)

    out_path = out_path or (append_to if append_to else path.with_suffix(".xlsx"))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out_path)
    return out_path


def build_parser() -> argparse.ArgumentParser:
    """Build the command-line parser.

    入力 TSV、テンプレート、マッピング、出力先を受け取る。
    """
    parser = argparse.ArgumentParser(description="Convert TSV to a custom Excel workbook.")
    parser.add_argument("path", type=Path, help="Input menu.tsv path.")
    parser.add_argument("--template", type=Path, help="Coach-provided template workbook.")
    parser.add_argument("--mapping", type=Path, default=repo_root() / "data" / "excel-template-mapping.json")
    parser.add_argument("--out", type=Path, help="Output .xlsx path.")
    parser.add_argument("--append-to", type=Path, help="Append a new sheet to an existing workbook (e.g. SPA.xlsx history).")
    parser.add_argument("--new-sheet-name", help="Name for the newly added sheet (defaults to the TSV Date meta).")
    return parser


def main(argv: list[str] | None = None) -> int:
    """Run the custom Excel export CLI.

    成功時は出力パスを表示する。
    """
    args = build_parser().parse_args(argv)
    try:
        out_path = convert_to_excel(
            args.path,
            args.template,
            args.mapping,
            args.out,
            append_to=args.append_to,
            new_sheet_name=args.new_sheet_name,
        )
    except Exception as exc:  # noqa: BLE001 - CLI boundary
        print(f"error: {exc}", file=sys.stderr)
        return 1
    print(out_path)
    return 0


if __name__ == "__main__":
    sys.exit(main())
