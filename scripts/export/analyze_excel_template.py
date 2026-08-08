"""Suggest a mapping for a coach Excel template.

Excel テンプレートのタイトルセル、日付セル、ヘッダ行、本文開始行、TOTAL 行、
列対応を推定し、`excel-template-mapping.json` の候補を出力する。
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any


HEADER_ALIASES = {
    "category": "category",
    "times": "times",
    "distance": "distance",
    "set": "set",
    "cycle": "cycle",
    "description": "description",
    "gears": "gears",
    "subtotal": "subtotal",
}


def import_openpyxl() -> Any:
    """Import openpyxl with a helpful error.

    依存がない場合はインストール方法を含む例外にする。
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required. Install with: python -m pip install openpyxl") from exc
    return openpyxl


def normalize(value: Any) -> str:
    """Normalize a cell value for matching.

    空白や記号を削り、小文字化する。
    """
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def cell_text(value: Any) -> str:
    """Convert a cell value to text.

    None は空文字として扱う。
    """
    return str(value).strip() if value is not None else ""


def first_non_empty_in_row(ws: Any, row_number: int) -> str | None:
    """Find the first non-empty cell in a row.

    タイトル候補として 1 行目を調べる。
    """
    for cell in ws[row_number]:
        if cell_text(cell.value):
            return cell.coordinate
    return None


def find_date_cell(ws: Any) -> str | None:
    """Find a likely date cell.

    Date ラベルまたは日付らしい値を含むセルを探す。
    """
    for row in ws.iter_rows():
        for cell in row:
            value = cell_text(cell.value)
            if not value:
                continue
            if normalize(value) == "date" or "日付" in value:
                return cell.coordinate
            if re.search(r"\d{4}[-/.]\d{1,2}[-/.]\d{1,2}", value):
                return cell.coordinate
    return None


def find_header_row(ws: Any) -> tuple[int | None, dict[str, str]]:
    """Find the likely header row and column mapping.

    Category/Distance などを含む行をメニュー表ヘッダとして推定する。
    """
    for row in ws.iter_rows():
        columns: dict[str, str] = {}
        for cell in row:
            key = HEADER_ALIASES.get(normalize(cell.value))
            if key:
                columns[key] = cell.column_letter
        if len(columns) >= 3:
            return row[0].row, columns
    return None, {}


def find_total_row(ws: Any) -> int | None:
    """Find the likely TOTAL row.

    セル値が TOTAL の行番号を返す。
    """
    for row in ws.iter_rows():
        if any(cell_text(cell.value).upper() == "TOTAL" for cell in row):
            return row[0].row
    return None


def analyze_template(path: Path) -> dict[str, Any]:
    """Analyze an Excel workbook and propose a mapping.

    先頭シートを対象にセル位置の候補を返す。
    """
    openpyxl = import_openpyxl()
    workbook = openpyxl.load_workbook(path, data_only=True)
    ws = workbook.active
    header_row, columns = find_header_row(ws)
    total_row = find_total_row(ws)
    return {
        "sheet_name": ws.title,
        "title": first_non_empty_in_row(ws, 1),
        "date": find_date_cell(ws),
        "header_row": header_row,
        "body_start_row": header_row + 1 if header_row else None,
        "body_end_row": total_row - 1 if total_row and header_row else None,
        "total_row": total_row,
        "columns": columns,
        "pace_table_start_row": None,
        "pace_columns": {},
        "source_template": str(path),
    }


def write_output(mapping: dict[str, Any], out_path: Path | None) -> None:
    """Write proposed mapping as JSON.

    出力先が指定されない場合は標準出力に出す。
    """
    payload = json.dumps(mapping, ensure_ascii=False, indent=2)
    if out_path:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(payload + "\n", encoding="utf-8")
    else:
        print(payload)


def build_parser() -> argparse.ArgumentParser:
    """Build the command-line parser.

    テンプレート Excel と任意の出力先を受け取る。
    """
    parser = argparse.ArgumentParser(description="Analyze an Excel template and suggest a mapping JSON.")
    parser.add_argument("path", type=Path, help="Input .xlsx template path.")
    parser.add_argument("--out", type=Path, help="Output mapping JSON file.")
    return parser


def main(argv: list[str] | None = None) -> int:
    """Run the template analyzer CLI.

    成功時は mapping JSON を出力する。
    """
    args = build_parser().parse_args(argv)
    try:
        mapping = analyze_template(args.path)
        write_output(mapping, args.out)
    except Exception as exc:  # noqa: BLE001 - CLI boundary
        print(f"error: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
