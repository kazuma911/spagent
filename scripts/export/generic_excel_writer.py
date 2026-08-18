"""Generic Excel writer using Excel Layout Descriptor v2.

Descriptor v2 (`data/excel-templates/<layout_id>.json`) と TSV から、コーチ提供の
Excel レイアウトに忠実な新シートを生成する。**特定コーチに依存しない汎用実装**。

対応要素 (descriptor 宣言で切替可能):

- ヘッダセル (team_name / datetime / equipment / facility / theme)
- Excel Table (ListObject) + 構造化参照数式 (`table[[#This Row],[Header]]`)
- セクションヘッダ行 (B 列のみに文字列、他は空 — descriptor で列変更可)
- 空行によるブロック区切り
- Description 列内の 【選手名】 個別注記 (regex パターンは descriptor で上書き可)
- TOTAL 行 (SUM − 個別 note 減算)

**数式 type セット** (4 種):
    - `product`                       — Times × Distance × Set 等の積式
    - `parse_cycle_to_hms`            — Cycle 文字列 (`1'30"`) → 0:MM:SS 変換
    - `cumulative_time`               — 累積時計 (L_prev + block_secs)
    - `sum_minus_individual_notes`    — 全体 SUM から個別 note 行を減算

**依存**: `openpyxl`
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shutil
import subprocess
import sys
from datetime import time
from pathlib import Path
from typing import Any

# Import style helpers from analyze package.
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "analyze"))
from _excel_style import apply_sheet_dimensions, apply_style_dict  # noqa: E402


def repo_root() -> Path:
    """Return repository root.

    スクリプトの場所からリポジトリルートを推定する。
    """
    return Path(__file__).resolve().parents[2]


def import_openpyxl() -> Any:
    """Import openpyxl with a helpful error.

    依存がない場合はインストール方法を含む例外にする。
    """
    try:
        import openpyxl
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise RuntimeError("openpyxl is required. Install with: python -m pip install openpyxl") from exc
    return openpyxl, Table, TableStyleInfo


def normalize_key(value: str) -> str:
    """Normalize a TSV comment / column key.

    メタ情報キーや列名を snake_case に統一する。
    """
    return re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")


def parse_tsv(path: Path) -> tuple[dict[str, str], list[dict[str, Any]]]:
    """Parse a menu TSV supporting section headers and blank separators.

    - `# key: value` 形式のメタ情報
    - `# <section-name>` 形式のセクションヘッダ行 (単独)
    - 空行はブロック区切り (row_kind=blank として保持)

    戻り値は `(meta, rows)` で、`rows` は `{row_kind, ...columns}` の並び。
    """
    meta: dict[str, str] = {}
    section_lines: list[tuple[str, str]] = []
    header_line: str | None = None
    body_lines: list[str] = []
    in_body = False

    with path.open("r", encoding="utf-8-sig", newline="") as file:
        for raw in file:
            line = raw.rstrip("\n")
            stripped = line.strip()
            if stripped.startswith("#"):
                text = stripped.lstrip("#").strip()
                if ":" in text and not text.startswith("["):
                    key, value = text.split(":", 1)
                    meta[normalize_key(key)] = value.strip()
                else:
                    section_lines.append(("section", text))
                    if in_body:
                        body_lines.append("__SECTION__" + text)
                continue
            if not in_body:
                if "\t" in line:
                    header_line = line
                    in_body = True
                continue
            body_lines.append(line)

    if not header_line:
        return meta, []

    headers = [normalize_key(cell) for cell in header_line.split("\t")]

    rows: list[dict[str, Any]] = []
    for line in body_lines:
        if line.startswith("__SECTION__"):
            rows.append({"row_kind": "section", "section_name": line[len("__SECTION__"):]})
            continue
        if not line.strip():
            rows.append({"row_kind": "blank"})
            continue
        cells = line.split("\t")
        row = {"row_kind": "data"}
        for idx, header in enumerate(headers):
            row[header] = cells[idx].strip() if idx < len(cells) else ""
        rows.append(row)

    return meta, rows


def load_descriptor(path: Path) -> dict[str, Any]:
    """Load Excel Layout Descriptor v2 JSON.

    バージョンチェックはしないが最低限のキーが揃っているか確認する。
    """
    with path.open("r", encoding="utf-8") as file:
        data = json.load(file)
    for key in ("layout_id", "header_cells", "table"):
        if key not in data:
            raise ValueError(f"descriptor missing required key: {key}")
    return data


def _col_letter(coord: str) -> str:
    """Extract column letters from a cell coordinate.

    `B12` → `B` のように英字部分だけ返す。
    """
    match = re.match(r"([A-Za-z]+)", coord)
    if not match:
        raise ValueError(f"invalid cell coordinate: {coord}")
    return match.group(1).upper()


def _int_or_none(value: str) -> int | None:
    """Parse an integer if possible.

    数値でなければ None を返す。
    """
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def _parse_time_seed(value: str) -> time | None:
    """Parse `HH:MM` or `H:MM` into a datetime.time.

    Descriptor の start seed を Excel の time 型として書き込む。
    """
    match = re.match(r"^\s*(\d{1,2}):(\d{2})\s*$", str(value or ""))
    if not match:
        return None
    hh, mm = int(match.group(1)), int(match.group(2))
    if 0 <= hh < 24 and 0 <= mm < 60:
        return time(hh, mm)
    return None


def _parse_time_string_formula(table_name: str, input_header: str, input_format: str = "cycle") -> str:
    """Return a formula converting a time string in the input column to 0:MM:SS.

    ``input_format``:
        - ``cycle`` : ``M'S"`` / ``M'`` / ``S"`` (Excel Table structured-reference dispatch).
        - ``pace``  : 同上 (Cycle と pace はどちらも M'S" 表記なので実装は共通).
        - ``hms``   : ``M:S`` or ``H:M:S`` (単純 TIMEVALUE).
    """
    ref = f"{table_name}[[#This Row],[{input_header}]]"
    if input_format == "hms":
        return f'=IFERROR(TIMEVALUE({ref}),0)'
    both = (
        f'IF(COUNTIF({ref},"*\'*""")>0,'
        f'_xlfn.TEXTJOIN(":",TRUE,0,'
        f'LEFT({ref},FIND("\'",{ref})-1),'
        f'MID({ref},FIND("\'",{ref})+1,FIND("""",{ref})-FIND("\'",{ref})-1)),'
    )
    only = (
        f'_xlfn.TEXTJOIN(":",TRUE,0,'
        f'IF(COUNTIF({ref},"*\'*")>0,LEFT({ref},FIND("\'",{ref})-1),0),'
        f'IF(COUNTIF({ref},"*""*")>0,LEFT({ref},FIND("""",{ref})-1),0)))'
    )
    return "=" + both + only


def _render_custom_expression(template: str, table_name: str, labels: dict[str, str], row_index: int, seed_row: int, formula_cols: dict[str, str]) -> str:
    """Substitute placeholders in a custom formula template.

    サポートするプレースホルダ:
        - ``{T}``                       — Table 名
        - ``{col.<key>}``               — 該当 key のヘッダラベル (構造化参照用)
        - ``{colref.<key>}``            — 該当 key の完全構造化参照 ``T[[#This Row],[label]]``
        - ``{seed_row}``                — seed cell の行番号
        - ``{prev_row}``                — このセルの一つ上の行番号 (境界時は seed_row)
        - ``{cellref.<key>.prev}``      — 該当 key の直前セル参照 (例: ``L11``)
    """
    prev = row_index - 1 if row_index > seed_row + 1 else seed_row
    out = template
    out = out.replace("{T}", table_name)
    out = out.replace("{seed_row}", str(seed_row))
    out = out.replace("{prev_row}", str(prev))
    for key, label in labels.items():
        out = out.replace("{col." + key + "}", label)
        out = out.replace("{colref." + key + "}", f"{table_name}[[#This Row],[{label}]]")
    for key, col_letter in formula_cols.items():
        out = out.replace("{cellref." + key + ".prev}", f"{col_letter}{prev}")
    return out


def build_row_formula(
    spec: dict[str, Any],
    row_index: int,
    table_name: str,
    labels: dict[str, str],
    seed_row: int,
    formula_cols: dict[str, str],
) -> str:
    """Dispatch per-row formula generation on ``spec['type']``.

    サポート type: ``product`` / ``parse_time_string`` / ``cumulative_from_column`` /
    ``custom_expression``. TOTAL 行向け type (sum_column / sum_minus_dynamic_ranges) は
    ここでは扱わず ``build_total_formula`` へ委譲する。
    """
    formula_type = spec.get("type")
    if formula_type == "product":
        operands = spec.get("operands") or []
        parts = [f"{table_name}[[#This Row],[{labels[op]}]]" for op in operands if op in labels]
        return "=" + "*".join(parts) if parts else ""
    if formula_type == "parse_time_string":
        input_key = spec.get("input")
        if input_key not in labels:
            return ""
        return _parse_time_string_formula(table_name, labels[input_key], spec.get("input_format", "cycle"))
    if formula_type == "cumulative_from_column":
        add_key = spec.get("add_column")
        col_letter = spec.get("col")
        if not add_key or add_key not in labels or not col_letter:
            return ""
        prev = row_index - 1 if row_index > seed_row + 1 else seed_row
        return f"={col_letter}{prev}+{table_name}[[#This Row],[{labels[add_key]}]]"
    if formula_type == "custom_expression":
        template = spec.get("template", "")
        return _render_custom_expression(template, table_name, labels, row_index, seed_row, formula_cols)
    return ""


def _consolidate_ranges(col_letter: str, sorted_rows: list[int]) -> list[str]:
    """Group consecutive row indices into contiguous ranges for SUM subtraction.

    ``[33, 34, 35, 37]`` → ``["I33:I35", "I37:I37"]`` のように圧縮する。
    """
    if not sorted_rows:
        return []
    ranges: list[str] = []
    start = sorted_rows[0]
    end = start
    for r in sorted_rows[1:]:
        if r == end + 1:
            end = r
        else:
            ranges.append(f"{col_letter}{start}:{col_letter}{end}")
            start = r
            end = r
    ranges.append(f"{col_letter}{start}:{col_letter}{end}")
    return ranges


def build_total_formula(
    spec: dict[str, Any],
    table_name: str,
    labels: dict[str, str],
    tagged_rows: dict[str, list[int]],
) -> str:
    """Dispatch TOTAL-row formula generation.

    サポート type:
        - ``sum_column``               — ``=SUM(T[[label]])``
        - ``sum_minus_dynamic_ranges`` — sum_column から filter マッチ行を減算
        - ``custom_expression``        — テンプレート文字列を使用 (limited)

    ``tagged_rows`` は動的フィルタごとの該当行番号リスト
    (``{"individual_note": [33, 35, 36, 38], ...}``)。
    """
    formula_type = spec.get("type")
    of_key = spec.get("of") or spec.get("input")
    if of_key not in labels:
        return ""
    col_letter = spec.get("col")
    base_sum = f"=SUM({table_name}[[{labels[of_key]}]])"
    if formula_type == "sum_column":
        return base_sum
    if formula_type == "sum_minus_dynamic_ranges":
        rows: list[int] = []
        for filter_name in spec.get("subtract_filters", ["individual_note"]):
            rows.extend(tagged_rows.get(filter_name, []))
        rows = sorted(set(rows))
        ranges = _consolidate_ranges(col_letter, rows) if col_letter else []
        return base_sum + "".join(f"-SUM({rng})" for rng in ranges)
    if formula_type == "custom_expression":
        template = spec.get("template", "")
        # TOTAL 行では row_index/seed_row を意識しないので placeholder を空で埋める
        return _render_custom_expression(template, table_name, labels, 0, 0, {})
    return ""


def _write_header_cells(ws: Any, descriptor: dict[str, Any], meta: dict[str, str]) -> None:
    """Write meta values into header cells declared in the descriptor.

    Descriptor `header_cells` の各セルへ TSV meta を配置する。
    """
    header_cells = descriptor.get("header_cells", {})
    key_map = {
        "team_name": ["team_name", "team"],
        "datetime": ["datetime", "date_time", "date"],
        "equipment": ["equipment"],
        "facility": ["facility"],
        "theme": ["theme", "title"],
    }
    for key, spec in header_cells.items():
        cell = spec.get("cell") if isinstance(spec, dict) else spec
        if not cell:
            continue
        candidates = key_map.get(key, [key])
        value = next((meta[c] for c in candidates if meta.get(c)), "")
        prefix_regex = spec.get("prefix_regex") if isinstance(spec, dict) else None
        if key == "equipment" and value and prefix_regex and not re.search(prefix_regex, value):
            value = "Equipment: " + value
        ws[cell] = value


def _write_column_headers(ws: Any, descriptor: dict[str, Any]) -> list[tuple[str, str, str]]:
    """Write the table header row and return column info.

    戻り値は `[(key, col_letter, header_label), ...]` の順序リスト。
    """
    table = descriptor["table"]
    header_row = int(table["header_row"])
    columns = table.get("columns", {})
    order: list[tuple[str, str, str]] = []
    for key, spec in columns.items():
        col = spec["col"]
        label = spec.get("header_label", key)
        ws[f"{col}{header_row}"] = label
        order.append((key, col, label))
    return order


def _dedupe_labels(order: list[tuple[str, str, str]]) -> list[tuple[str, str, str]]:
    """Ensure column labels are unique for Excel Table structured references.

    ``header_label`` が空白のみで重複する場合、末尾に空白を追加してユニーク化する。
    Excel Table (ListObject) は同名ヘッダを許容しないため必須。
    """
    seen: dict[str, int] = {}
    out: list[tuple[str, str, str]] = []
    for key, col, label in order:
        base = label if label else " "
        candidate = base
        while candidate in seen:
            candidate = base + " " * (seen[base] + 1)
            seen[base] = seen.get(base, 0) + 1
        seen[candidate] = 1
        out.append((key, col, candidate))
    return out


def _row_formula_specs(descriptor: dict[str, Any]) -> list[tuple[str, dict[str, Any]]]:
    """Return per-row formula specs (excluding TOTAL-row-only types), in declared order.

    ``sum_column`` / ``sum_minus_dynamic_ranges`` は TOTAL 行専用なので除外。
    """
    row_types = {"product", "parse_time_string", "cumulative_from_column", "custom_expression"}
    formulas = descriptor.get("formulas", {}) or {}
    return [
        (name, spec) for name, spec in formulas.items()
        if isinstance(spec, dict) and spec.get("type") in row_types
    ]


def _total_formula_specs(descriptor: dict[str, Any]) -> list[tuple[str, dict[str, Any]]]:
    """Return TOTAL-row formula specs (only sum-based types)."""
    total_types = {"sum_column", "sum_minus_dynamic_ranges"}
    formulas = descriptor.get("formulas", {}) or {}
    return [
        (name, spec) for name, spec in formulas.items()
        if isinstance(spec, dict) and spec.get("type") in total_types
    ]


def _write_literal_data(
    ws: Any,
    row_index: int,
    row: dict[str, Any],
    order: list[tuple[str, str, str]],
    formula_keys: set[str],
) -> None:
    """Write literal (non-formula) column values for one data row.

    数値型宣言のある列は int 化を試み、失敗時は文字列で書く。
    """
    for key, col, _ in order:
        if key in formula_keys:
            continue
        cell = f"{col}{row_index}"
        value = row.get(key, "")
        if key in {"times", "distance", "set"}:
            n = _int_or_none(value)
            if n is not None:
                ws[cell] = n
                continue
            if key == "set" and (value == "" or value is None):
                ws[cell] = 1
                continue
        ws[cell] = value


def _write_data_row(
    ws: Any,
    row_index: int,
    row: dict[str, Any],
    order: list[tuple[str, str, str]],
    table_name: str,
    labels: dict[str, str],
    seed_row: int,
    row_specs: list[tuple[str, dict[str, Any]]],
    formula_cols: dict[str, str],
) -> None:
    """Write one data row: literal fields + descriptor-driven formulas.

    Formula 順序は descriptor 記述順を維持する (依存解決のため cycle_secs → block_secs
    → elapsed のような順に並べる想定)。
    """
    formula_keys = {name for name, _ in row_specs}
    _write_literal_data(ws, row_index, row, order, formula_keys)
    for name, spec in row_specs:
        col_letter = spec.get("col")
        if not col_letter:
            continue
        expression = build_row_formula(spec, row_index, table_name, labels, seed_row, formula_cols)
        if expression:
            ws[f"{col_letter}{row_index}"] = expression


def _write_continuation_row(
    ws: Any,
    row_index: int,
    seed_row: int,
    row_specs: list[tuple[str, dict[str, Any]]],
    formula_cols: dict[str, str],
    label_or_section: tuple[str, str] | None,
) -> None:
    """Handle blank / section rows honoring per-formula ``on_blank_row``.

    ``continue_prev`` (既定): ``={col}{prev}`` を書き累積を継続する。
    ``skip``               : 何も書かない。

    ``label_or_section`` = ``(col_letter, text)`` があればセクション名を書き込む。
    """
    if label_or_section is not None:
        col, text_ = label_or_section
        ws[f"{col}{row_index}"] = text_
    for name, spec in row_specs:
        mode = spec.get("on_blank_row", "continue_prev" if spec.get("type") == "cumulative_from_column" else "skip")
        if mode == "skip":
            continue
        col_letter = spec.get("col")
        if not col_letter:
            continue
        prev = row_index - 1 if row_index > seed_row + 1 else seed_row
        ws[f"{col_letter}{row_index}"] = f"={col_letter}{prev}"


def _write_total_row(
    ws: Any,
    row_index: int,
    table_name: str,
    labels: dict[str, str],
    tagged_rows: dict[str, list[int]],
    descriptor: dict[str, Any],
) -> None:
    """Write the TOTAL row per descriptor total-formula specs.

    サポート type: ``sum_column`` / ``sum_minus_dynamic_ranges`` / ``custom_expression``.
    """
    total_specs = _total_formula_specs(descriptor)
    total_marker = descriptor.get("formulas", {}).get("total_marker", {})
    marker_col = total_marker.get("column")
    marker_text = total_marker.get("text", "Total")
    if not marker_col:
        # 互換: 最初の total spec の row_marker_column/text
        for _, spec in total_specs:
            marker_col = spec.get("row_marker_column", marker_col)
            marker_text = spec.get("row_marker_text", marker_text)
            if marker_col:
                break
    if marker_col:
        ws[f"{marker_col}{row_index}"] = marker_text
    for _, spec in total_specs:
        col_letter = spec.get("col")
        if not col_letter:
            continue
        expression = build_total_formula(spec, table_name, labels, tagged_rows)
        if expression:
            ws[f"{col_letter}{row_index}"] = expression


def _apply_descriptor_styles(
    ws: Any,
    descriptor: dict[str, Any],
    header_row: int,
    body_start: int,
    last_row_exclusive: int,
    order: list[tuple[str, str, str]],
) -> None:
    """Apply Descriptor v2 `styles` section to the freshly written sheet.

    - header_cells: 各名前付きセル
    - table_header_row: header 行全列
    - body_row_default.by_column: body 行の列ごと
    - section_header_row / total_row: 該当行 (簡易ヒューリスティック)
    - column_widths / row_heights
    """
    styles = descriptor.get("styles") or {}
    if not styles:
        return

    header_cells_conf = descriptor.get("header_cells") or {}
    for name, style in (styles.get("header_cells") or {}).items():
        target_cell = (header_cells_conf.get(name) or {}).get("cell")
        if target_cell:
            apply_style_dict(ws[target_cell], style)

    header_style = styles.get("table_header_row")
    if header_style:
        for _, col_letter, _ in order:
            apply_style_dict(ws[f"{col_letter}{header_row}"], header_style)

    body_by_col = ((styles.get("body_row_default") or {}).get("by_column") or {})
    if body_by_col:
        canonical_to_letter: dict[str, str] = {}
        for canonical, col_letter, _ in order:
            canonical_to_letter[canonical] = col_letter
        for canonical, spec in (descriptor.get("formulas") or {}).items():
            if isinstance(spec, dict) and spec.get("col"):
                canonical_to_letter[canonical] = spec["col"]
        for r in range(body_start, last_row_exclusive):
            for canonical, col_letter in canonical_to_letter.items():
                style = body_by_col.get(canonical)
                if style:
                    apply_style_dict(ws[f"{col_letter}{r}"], style)

    section_style = styles.get("section_header_row")
    if section_style:
        category_col = descriptor["table"]["columns"]["category"]["col"]
        other_letters = [c for _, c, _ in order if c != category_col]
        for r in range(body_start, last_row_exclusive):
            first_val = ws[f"{category_col}{r}"].value
            if isinstance(first_val, str) and first_val.strip() and all(
                ws[f"{c}{r}"].value in (None, "") for c in other_letters
            ):
                apply_style_dict(ws[f"{category_col}{r}"], section_style)

    total_style = styles.get("total_row")
    total_marker = ((descriptor.get("formulas") or {}).get("total_marker") or {})
    total_col = total_marker.get("column")
    if total_style and total_col:
        for r in range(last_row_exclusive - 1, body_start - 1, -1):
            v = ws[f"{total_col}{r}"].value
            if isinstance(v, str) and v.strip().lower() in {"total", "合計", "計"}:
                for _, c, _ in order:
                    apply_style_dict(ws[f"{c}{r}"], total_style)
                break

    apply_sheet_dimensions(
        ws,
        {
            "column_widths": styles.get("column_widths") or {},
            "row_heights": styles.get("row_heights") or {},
        },
    )


def _register_table(workbook: Any, ws: Any, order: list[tuple[str, str, str]], header_row: int, last_row: int, descriptor: dict[str, Any], Table: Any, TableStyleInfo: Any) -> str:
    """Register a Table (ListObject) covering header_row..last_row.

    Excel Table 名はワークブック内でユニークでなければならないため、
    `<prefix><sheet-index>` の形で一意化する。
    """
    prefix = descriptor.get("table", {}).get("table_name_prefix", "Menu")
    style_name = descriptor.get("table", {}).get("table_style", "TableStyleMedium2")
    columns = [col for _, col, _ in order]
    first_col = columns[0]
    last_col = columns[-1]
    ref = f"{first_col}{header_row}:{last_col}{last_row}"
    existing = {tbl for sh in workbook.worksheets for tbl in getattr(sh, "tables", {})}
    idx = len(existing) + 1
    while f"{prefix}{idx}" in existing:
        idx += 1
    table_name = f"{prefix}{idx}"
    table = Table(displayName=table_name, name=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    return table_name


def write_workbook(
    tsv_path: Path,
    descriptor_path: Path,
    out_path: Path,
    *,
    append_to: Path | None = None,
    new_sheet_name: str | None = None,
    start_time_override: str | None = None,
) -> Path:
    """Write a workbook honoring the descriptor.

    `append_to` を指定すると既存 xlsx に新シートを追加する。
    """
    openpyxl, Table, TableStyleInfo = import_openpyxl()
    descriptor = load_descriptor(descriptor_path)
    meta, rows = parse_tsv(tsv_path)

    template_source = descriptor.get("template_source") or {}
    template_xlsx = _resolve_template_xlsx(template_source, descriptor_path)
    template_mode = template_xlsx is not None and not append_to

    if append_to:
        workbook = openpyxl.load_workbook(append_to)
        sheet_name = new_sheet_name or meta.get("sheet_name") or meta.get("date") or tsv_path.stem
        while sheet_name in workbook.sheetnames:
            sheet_name = f"{sheet_name}-2"
        ws = workbook.create_sheet(title=sheet_name)
        _write_header_cells(ws, descriptor, meta)
        _write_column_headers_raw = True
    elif template_mode:
        sheet_name = new_sheet_name or meta.get("sheet_name") or meta.get("date") or tsv_path.stem
        workbook, ws = _prepare_from_template(
            template_xlsx, template_source.get("sheet_name"), out_path, sheet_name, openpyxl
        )
        _clear_body_and_tables(ws, int(descriptor["table"]["header_row"]))
        _write_header_cells(ws, descriptor, meta)
        _write_column_headers_raw = False
    else:
        workbook = openpyxl.Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        sheet_name = new_sheet_name or meta.get("sheet_name") or meta.get("date") or tsv_path.stem
        while sheet_name in workbook.sheetnames:
            sheet_name = f"{sheet_name}-2"
        ws = workbook.create_sheet(title=sheet_name)
        _write_header_cells(ws, descriptor, meta)
        _write_column_headers_raw = True

    if _write_column_headers_raw:
        order = _write_column_headers(ws, descriptor)
        # Non-template mode: also write header labels for formula columns.
        header_row_early = int(descriptor["table"]["header_row"])
        extended = _order_from_descriptor(descriptor)
        for key, col, label in extended:
            if not any(existing[1] == col for existing in order):
                ws[f"{col}{header_row_early}"] = label
                order.append((key, col, label))
        order.sort(key=lambda t: _column_index(t[1]))
        order = _dedupe_labels(order)
    else:
        # Template mode: read actual header text from the ws so structured refs match.
        order = _order_from_descriptor(descriptor)
        header_row_early = int(descriptor["table"]["header_row"])
        rewritten: list[tuple[str, str, str]] = []
        for key, col, _label in order:
            actual = ws[f"{col}{header_row_early}"].value
            rewritten.append((key, col, actual if actual is not None else _label))
        order = rewritten
    labels = {key: label for key, _, label in order}

    header_row = int(descriptor["table"]["header_row"])
    body_start = int(descriptor["table"]["body_start_row"])
    seed_spec = descriptor.get("start_time_seed", {}) or {}
    seed_cell = seed_spec.get("cell", "L6")
    seed_row = int(re.search(r"\d+", seed_cell).group(0))
    seed_value = start_time_override or meta.get("start_time") or seed_spec.get("example_value")
    seed_time = _parse_time_seed(seed_value) if seed_value else None
    if seed_time is not None:
        ws[seed_cell] = seed_time

    individual_re = re.compile(
        (descriptor.get("individual_note", {}) or {}).get("prefix_regex", r"^【([^】]+)】")
    )
    category_col = _col_letter(descriptor["table"]["columns"]["category"]["col"])

    row_specs = _row_formula_specs(descriptor)
    formula_cols = {name: spec.get("col") for name, spec in row_specs if spec.get("col")}

    # 1st pass: count rows to compute the Table range
    current_row = body_start
    last_data_row = body_start - 1
    for row in rows:
        kind = row.get("row_kind", "data")
        if kind in ("section", "blank"):
            current_row += 1
            continue
        current_row += 1
        last_data_row = current_row - 1

    if last_data_row < body_start:
        last_data_row = body_start

    table_name = _register_table(workbook, ws, order, header_row, last_data_row, descriptor, Table, TableStyleInfo)

    # 2nd pass: write rows with formulas resolved against the actual table name
    current_row = body_start
    tagged_rows: dict[str, list[int]] = {"individual_note": []}
    for row in rows:
        kind = row.get("row_kind", "data")
        if kind == "section":
            _write_continuation_row(
                ws, current_row, seed_row, row_specs, formula_cols,
                (category_col, row.get("section_name", "")),
            )
            current_row += 1
            continue
        if kind == "blank":
            _write_continuation_row(ws, current_row, seed_row, row_specs, formula_cols, None)
            current_row += 1
            continue
        _write_data_row(ws, current_row, row, order, table_name, labels, seed_row, row_specs, formula_cols)
        desc_value = str(row.get("description", "") or "")
        if individual_re.match(desc_value):
            tagged_rows["individual_note"].append(current_row)
        current_row += 1

    # In template mode, retarget the TOTAL row to the original template row so its
    # formatting (borders, merged marker) is preserved. Then trim empty rows between
    # data and TOTAL to eliminate whitespace.
    if template_mode:
        original_total_row = (descriptor.get("formulas", {}).get("total_marker") or {}).get("template_row")
        if original_total_row and current_row < original_total_row:
            gap = original_total_row - current_row
            try:
                _shift_up_delete(ws, current_row, gap)
                new_total_row = current_row
            except Exception as e:
                print(f"warn: _shift_up_delete(gap) failed: {type(e).__name__}: {e}", file=sys.stderr)
                new_total_row = original_total_row
        else:
            new_total_row = current_row
        _write_total_row(ws, new_total_row, table_name, labels, tagged_rows, descriptor)
        last_written_row = new_total_row
        try:
            _prune_below(ws, last_written_row)
        except Exception as e:
            print(f"warn: _prune_below failed: {type(e).__name__}: {e}", file=sys.stderr)
    else:
        _write_total_row(ws, current_row, table_name, labels, tagged_rows, descriptor)
        last_written_row = current_row
        _apply_descriptor_styles(ws, descriptor, header_row, body_start, current_row, order)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    if append_to:
        workbook.save(append_to)
        return append_to
    workbook.save(out_path)
    return out_path


def _shift_up_delete(ws: Any, start_row: int, count: int) -> None:
    """Delete `count` rows starting at `start_row`, shifting merged ranges accordingly.

    ``openpyxl.Worksheet.delete_rows`` removes cell values but leaves merged-cell
    ranges pinned to their old row numbers, causing Excel to still render whitespace
    up to the pinned row. This helper first records merged ranges, drops the affected
    rows, then re-registers each range at its shifted position.
    """
    from openpyxl.worksheet.cell_range import CellRange

    old_ranges = [(str(r), CellRange(str(r))) for r in list(ws.merged_cells.ranges)]
    for name, _ in old_ranges:
        ws.unmerge_cells(name)

    ws.delete_rows(start_row, count)

    for _, rng in old_ranges:
        if rng.max_row < start_row:
            ws.merge_cells(str(rng))
        elif rng.min_row >= start_row + count:
            new_range = CellRange(min_col=rng.min_col, min_row=rng.min_row - count, max_col=rng.max_col, max_row=rng.max_row - count)
            ws.merge_cells(str(new_range))
        elif rng.min_row >= start_row and rng.max_row < start_row + count:
            continue
        else:
            new_min = rng.min_row if rng.min_row < start_row else start_row
            new_max = rng.max_row - count if rng.max_row >= start_row + count else start_row - 1
            if new_min <= new_max:
                new_range = CellRange(min_col=rng.min_col, min_row=new_min, max_col=rng.max_col, max_row=new_max)
                ws.merge_cells(str(new_range))


def _prune_below(ws: Any, last_row: int) -> None:
    """Remove trailing empty rows below `last_row` and any merged-cell debris."""
    from openpyxl.worksheet.cell_range import CellRange

    dead_ranges = [str(r) for r in list(ws.merged_cells.ranges) if CellRange(str(r)).min_row > last_row]
    for name in dead_ranges:
        ws.unmerge_cells(name)
    max_row = ws.max_row or last_row
    if max_row > last_row:
        ws.delete_rows(last_row + 1, max_row - last_row)
    for rn in list(ws.row_dimensions.keys()):
        if isinstance(rn, int) and rn > last_row:
            del ws.row_dimensions[rn]


def _resolve_template_xlsx(template_source: dict[str, Any], descriptor_path: Path) -> Path | None:
    """Locate the reference xlsx if available on disk.

    Descriptor は相対パスでも絶対パスでも受け付ける。見つからなければ None (styles 辞書へフォールバック)。
    """
    if not template_source:
        return None
    raw = template_source.get("xlsx_path")
    if not raw:
        return None
    for candidate in (Path(raw), Path(raw).expanduser(), (descriptor_path.parent / raw).resolve(), (repo_root() / raw).resolve()):
        if candidate.exists() and candidate.is_file():
            return candidate
    return None


def _prepare_from_template(template_xlsx: Path, template_sheet: str | None, out_path: Path, sheet_name: str, openpyxl: Any) -> tuple[Any, Any]:
    """Copy the template xlsx to `out_path`, keep only the reference sheet, and rename it.

    Merged cells / conditional formatting / borders / fonts / column widths / images
    はすべて維持される。data 部だけ後段で書き換える。
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_xlsx, out_path)
    workbook = openpyxl.load_workbook(out_path)
    target = template_sheet or workbook.sheetnames[-1]
    if target not in workbook.sheetnames:
        target = workbook.sheetnames[-1]
    for name in list(workbook.sheetnames):
        if name != target:
            del workbook[name]
    ws = workbook[target]
    if sheet_name and sheet_name != target:
        ws.title = sheet_name
    return workbook, ws


def _clear_body_and_tables(ws: Any, header_row: int) -> None:
    """Clear cell values below the header row (styles kept) and drop existing Tables.

    header 行より上 (team/facility/theme/equipment) は温存。header 行以下の値をクリア。
    Table (ListObject) は範囲が古いままだと Excel が壊れる可能性があるので削除。
    """
    for tbl_name in list((ws.tables or {}).keys()):
        del ws.tables[tbl_name]
    max_row = ws.max_row or header_row + 1
    max_col = ws.max_column or 12
    for r in range(header_row + 1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                cell.value = None


def _order_from_descriptor(descriptor: dict[str, Any]) -> list[tuple[str, str, str]]:
    """Rebuild the column order list without touching existing header cells.

    Template mode で header 行は既にオリジナルからコピー済みなので上書きしない。
    formula 列 (subtotal_m / cycle_secs / block_secs / elapsed) も Table 範囲に含める必要があるため
    ここで統合する。
    """
    order: list[tuple[str, str, str]] = []
    columns = descriptor.get("table", {}).get("columns", {})
    for key, spec in columns.items():
        col_letter = spec.get("col")
        if not col_letter:
            continue
        label = spec.get("header_label") or key
        order.append((key, col_letter, label))
    # Extend with formula columns so the Excel Table spans them too.
    formulas = descriptor.get("formulas", {}) or {}
    for name, spec in formulas.items():
        if not isinstance(spec, dict):
            continue
        col_letter = spec.get("col")
        if not col_letter or name in {"total_marker", "total_distance"}:
            continue
        if any(existing_col == col_letter for _, existing_col, _ in order):
            continue
        label = spec.get("header_label") or name
        order.append((name, col_letter, label))
    order.sort(key=lambda t: _column_index(t[1]))
    return order


def _column_index(letter: str) -> int:
    """Excel column letter → 0-based index."""
    idx = 0
    for ch in letter:
        idx = idx * 26 + (ord(ch.upper()) - ord("A") + 1)
    return idx - 1


def open_file_with_default_app(path: Path) -> bool:
    """OS 既定アプリでファイルを開く。

    Windows: os.startfile / macOS: `open` / Linux: `xdg-open`。
    成功で True、失敗で False。
    """
    p = Path(path).resolve()
    if not p.exists():
        return False
    try:
        if sys.platform.startswith("win"):
            os.startfile(str(p))  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(p)], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        else:
            subprocess.Popen(["xdg-open", str(p)], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        return True
    except Exception:
        return False


def build_parser() -> argparse.ArgumentParser:
    """Build the CLI parser.

    入力 TSV、Descriptor JSON、出力先、append_to、sheet 名を受け取る。
    """
    parser = argparse.ArgumentParser(description="Generate an Excel workbook using a Layout Descriptor v2.")
    parser.add_argument("tsv", type=Path, help="Input menu.tsv path.")
    parser.add_argument("--descriptor", type=Path, required=True, help="Layout descriptor v2 JSON.")
    parser.add_argument("--out", type=Path, help="Output .xlsx path (ignored if --append-to is set).")
    parser.add_argument("--append-to", type=Path, help="Append a new sheet to an existing workbook.")
    parser.add_argument("--new-sheet-name", help="Sheet name for the new sheet.")
    parser.add_argument("--start-time", help="Override start time (HH:MM) for the seed cell.")
    parser.add_argument("--open", dest="open_after", action="store_true", default=True, help="Open the output file with the OS default app (default: on).")
    parser.add_argument("--no-open", dest="open_after", action="store_false", help="Do not open the output file after generation.")
    return parser


def main(argv: list[str] | None = None) -> int:
    """CLI entrypoint.

    成功時は出力パスを stdout に出す。
    """
    args = build_parser().parse_args(argv)
    try:
        result = write_workbook(
            args.tsv,
            args.descriptor,
            args.out or args.tsv.with_suffix(".xlsx"),
            append_to=args.append_to,
            new_sheet_name=args.new_sheet_name,
            start_time_override=args.start_time,
        )
    except Exception as exc:  # noqa: BLE001 - CLI boundary
        print(f"error: {exc}", file=sys.stderr)
        return 1
    print(result)
    if args.open_after:
        target = Path(args.append_to) if args.append_to else Path(result)
        if open_file_with_default_app(target):
            print(f"opened: {target}", file=sys.stderr)
        else:
            print(f"warning: could not open {target}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
