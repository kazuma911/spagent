"""Style capture and reapply helpers for Excel Layout Descriptor v2.

**扱う要素**:

- Font (name, size, bold, italic, color)
- PatternFill (patternType, fgColor)
- Alignment (horizontal, vertical, wrap_text)
- Border (left/right/top/bottom style + color)
- number_format
- Column widths / row heights

**シリアライズ形式**: JSON dict にまとめて Descriptor に埋め込み可能。
Reader 側は openpyxl オブジェクトから `cell_style_to_dict()` で読み取り、
Writer 側は `apply_style_dict()` でセルに戻す。

**依存**: `openpyxl`
"""

from __future__ import annotations

from typing import Any


def _color_to_str(color: Any) -> str | None:
    """Return an ARGB / theme color as string, or None."""
    if color is None:
        return None
    try:
        if getattr(color, "type", None) == "rgb" and color.rgb:
            return str(color.rgb)
        if getattr(color, "type", None) == "theme":
            return f"theme:{color.theme}" + (f":{color.tint}" if color.tint else "")
        if getattr(color, "type", None) == "indexed":
            return f"indexed:{color.indexed}"
    except Exception:
        return None
    return None


def font_to_dict(font: Any) -> dict[str, Any]:
    """Serialize an openpyxl Font."""
    if font is None:
        return {}
    d: dict[str, Any] = {}
    if font.name:
        d["name"] = font.name
    if font.size:
        d["size"] = float(font.size)
    if font.bold:
        d["bold"] = True
    if font.italic:
        d["italic"] = True
    color = _color_to_str(font.color)
    if color:
        d["color"] = color
    return d


def fill_to_dict(fill: Any) -> dict[str, Any]:
    """Serialize an openpyxl PatternFill (returns empty dict for no fill)."""
    if fill is None:
        return {}
    try:
        pattern = getattr(fill, "patternType", None)
        if not pattern:
            return {}
        d: dict[str, Any] = {"patternType": pattern}
        fg = _color_to_str(getattr(fill, "fgColor", None))
        if fg:
            d["fgColor"] = fg
        return d
    except Exception:
        return {}


def alignment_to_dict(align: Any) -> dict[str, Any]:
    """Serialize an openpyxl Alignment."""
    if align is None:
        return {}
    d: dict[str, Any] = {}
    if align.horizontal:
        d["horizontal"] = align.horizontal
    if align.vertical:
        d["vertical"] = align.vertical
    if align.wrap_text:
        d["wrap_text"] = True
    return d


def _side_to_dict(side: Any) -> dict[str, Any]:
    """Serialize one Border side (openpyxl Side)."""
    if side is None or not getattr(side, "style", None):
        return {}
    d: dict[str, Any] = {"style": side.style}
    color = _color_to_str(getattr(side, "color", None))
    if color:
        d["color"] = color
    return d


def border_to_dict(border: Any) -> dict[str, Any]:
    """Serialize an openpyxl Border."""
    if border is None:
        return {}
    d: dict[str, Any] = {}
    for edge in ("left", "right", "top", "bottom"):
        side_d = _side_to_dict(getattr(border, edge, None))
        if side_d:
            d[edge] = side_d
    return d


def cell_style_to_dict(cell: Any) -> dict[str, Any]:
    """Serialize a single cell's styling to a plain dict."""
    d: dict[str, Any] = {}
    font = font_to_dict(cell.font)
    if font:
        d["font"] = font
    fill = fill_to_dict(cell.fill)
    if fill:
        d["fill"] = fill
    alignment = alignment_to_dict(cell.alignment)
    if alignment:
        d["alignment"] = alignment
    border = border_to_dict(cell.border)
    if border:
        d["border"] = border
    if cell.number_format and cell.number_format != "General":
        d["number_format"] = cell.number_format
    return d


def _resolve_color(value: str | None) -> Any | None:
    """Turn a serialized color back into an openpyxl Color."""
    from openpyxl.styles import Color

    if not value:
        return None
    if value.startswith("theme:"):
        parts = value.split(":")
        theme = int(parts[1])
        tint = float(parts[2]) if len(parts) > 2 else 0.0
        return Color(theme=theme, tint=tint)
    if value.startswith("indexed:"):
        return Color(indexed=int(value.split(":", 1)[1]))
    return Color(rgb=value)


def apply_style_dict(cell: Any, style: dict[str, Any] | None) -> None:
    """Apply a serialized style dict to a cell."""
    if not style:
        return
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    if "font" in style:
        f = style["font"]
        cell.font = Font(
            name=f.get("name"),
            size=f.get("size"),
            bold=f.get("bold", False),
            italic=f.get("italic", False),
            color=_resolve_color(f.get("color")),
        )
    if "fill" in style:
        fl = style["fill"]
        cell.fill = PatternFill(
            patternType=fl.get("patternType"),
            fgColor=_resolve_color(fl.get("fgColor")) or "FFFFFFFF",
        )
    if "alignment" in style:
        a = style["alignment"]
        cell.alignment = Alignment(
            horizontal=a.get("horizontal"),
            vertical=a.get("vertical"),
            wrap_text=a.get("wrap_text", False),
        )
    if "border" in style:
        b = style["border"]

        def _side(spec: dict[str, Any] | None) -> Side:
            if not spec:
                return Side()
            return Side(style=spec.get("style"), color=_resolve_color(spec.get("color")))

        cell.border = Border(
            left=_side(b.get("left")),
            right=_side(b.get("right")),
            top=_side(b.get("top")),
            bottom=_side(b.get("bottom")),
        )
    if "number_format" in style:
        cell.number_format = style["number_format"]


def sheet_dimensions_to_dict(ws: Any) -> dict[str, Any]:
    """Capture column widths and (sampled) row heights."""
    widths: dict[str, float] = {}
    for letter, dim in (ws.column_dimensions or {}).items():
        if dim and dim.width:
            widths[letter] = float(dim.width)
    heights: dict[str, float] = {}
    for rn, dim in (ws.row_dimensions or {}).items():
        if dim and dim.height:
            heights[str(rn)] = float(dim.height)
    return {"column_widths": widths, "row_heights": heights}


def apply_sheet_dimensions(ws: Any, dims: dict[str, Any] | None) -> None:
    """Restore column widths and row heights."""
    if not dims:
        return
    for letter, width in (dims.get("column_widths") or {}).items():
        ws.column_dimensions[letter].width = float(width)
    for rn_str, height in (dims.get("row_heights") or {}).items():
        try:
            rn = int(rn_str)
        except ValueError:
            continue
        ws.row_dimensions[rn].height = float(height)
