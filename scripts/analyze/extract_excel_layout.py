"""Extract Excel Layout Descriptor v2 from an arbitrary coach's xlsx.

**目的**: 手書きの descriptor を排除し、コーチが xlsx を投入するだけで
`data/excel-templates/<layout_id>.json` を自動生成する。

**検出パイプライン** (各段階に confidence スコア):

1. ヘッダセル分類 (team_name / datetime / equipment / facility / theme)
2. table header 行検出 (Category/Times/Distance/... と日本語別名)
3. 列マッピング
4. セクションヘッダパターン学習
5. 個別注記 bracket 検出 (【】/「」/[]/《》)
6. 数式パターン照合 → 数式 type 推定
7. TOTAL 行検出
8. start_time_seed 検出

confidence < 0.7 の項目は `warnings` として列挙し、コーチ確認を促す。
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from datetime import datetime, time
from pathlib import Path
from typing import Any

# Local import path resolution (script may run as `python scripts/analyze/...`)
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _excel_style import (  # noqa: E402
    cell_style_to_dict,
    sheet_dimensions_to_dict,
)


# =============================================================================
# 定数: 見出し語辞書 (英日両対応)
# =============================================================================

TABLE_HEADER_ALIASES = {
    # normalized_value: canonical_key
    "category": "category", "cat": "category", "種目": "category", "カテゴリ": "category", "区分": "category",
    "times": "times", "回数": "times", "本数": "times",
    "distance": "distance", "dist": "distance", "距離": "distance",
    "set": "set", "セット": "set", "組数": "set",
    "cycle": "cycle", "サイクル": "cycle", "cyc": "cycle",
    "description": "description", "desc": "description", "内容": "description", "説明": "description", "コメント": "description",
    "gears": "gears", "wgears": "gears", "gear": "gears", "器具": "gears", "wgear": "gears",
    "subtotal": "subtotal_m", "total": "subtotal_m", "小計": "subtotal_m",
}

# ヘッダ分類用の正規表現
DATETIME_RE = re.compile(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})(?:\s*\([^)]*\))?(?:\s+\d{1,2}:\d{2})?(?:\s*-\s*\d{1,2}:\d{2})?")
EQUIPMENT_PREFIX_RE = re.compile(r"^\s*equipment\s*[:：]", re.IGNORECASE)
EQUIPMENT_ITEMS_RE = re.compile(r"(fins?|paddle|board|buoy|snorkel|goggle|watch|stopwatch|band|器具|プドル|ビート板|パドル|フィン)", re.IGNORECASE)
FACILITY_RE = re.compile(r"(pool|プール|センター|アクアティクス|温水|ジム|体育館|LCM|SCM|\d{2,3}\s*m|レーン|lane)", re.IGNORECASE)
THEME_RE = re.compile(r"(Phase\s*[A-D]|D-\s*\d+|D\+\s*\d+|Taper|Threshold|Sprint|Recovery|Base|USRPT|Broken|Descending|Aerobic|Race\s*Pace|Endurance|IM\s*day|Kick\s*day|Drill\s*day|VO2|EN[123]|SP[12]|大会|試合|レース|基礎期|準備期|移行期|テーマ|theme|Trans|Build|フェーズ)", re.IGNORECASE)

# 個別注記 bracket 候補
BRACKET_PAIRS = [
    ("【", "】", r"^【([^】]+)】"),
    ("「", "」", r"^「([^」]+)」"),
    ("《", "》", r"^《([^》]+)》"),
    ("(", ")", r"^[(（]([^)）]+)[)）]"),
    ("[", "]", r"^\[([^\]]+)\]"),
]

# 数式パターン (Excel Table 構造化参照)
FORMULA_STRUCTURED_REF_RE = re.compile(r"\[\[#This Row\],\[([^\]]+)\]\]")
FORMULA_TABLE_REF_RE = re.compile(r"^=([A-Za-z_][A-Za-z0-9_]*)\[")
FORMULA_CELL_REF_RE = re.compile(r"([A-Z]+)(\d+)")
FORMULA_SUM_RE = re.compile(r"SUM\s*\(", re.IGNORECASE)
FORMULA_TEXTJOIN_RE = re.compile(r"TEXTJOIN", re.IGNORECASE)
FORMULA_FIND_QUOTE_RE = re.compile(r"FIND\s*\(\s*\"\'\"", re.IGNORECASE)

TOTAL_MARKER_RE = re.compile(r"^\s*(total|合計|総距離|計)\s*$", re.IGNORECASE)


# =============================================================================
# ユーティリティ
# =============================================================================


def import_openpyxl() -> Any:
    """Import openpyxl with a helpful error.

    依存がない場合はインストール方法を含む例外にする。
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required. Install with: python -m pip install openpyxl") from exc
    return openpyxl


def normalize(value: Any) -> str:
    """Normalize a cell value for header/key matching.

    英数字以外を全部落として小文字化。日本語文字はそのまま保持する。
    """
    if value is None:
        return ""
    text = str(value).strip().lower()
    # 半角記号・空白のみ削除、日本語は残す
    return re.sub(r"[\s_\-/\\.,、。]+", "", text)


def cell_text(value: Any) -> str:
    """Convert a cell value to display text.

    日付/時刻型は ISO/文字列に変換する。
    """
    if value is None:
        return ""
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).strip()


def column_letter(col_index: int) -> str:
    """Convert 1-based column index to Excel letter.

    ``1 → A``, ``27 → AA`` のような変換。
    """
    letters = ""
    n = col_index
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


# =============================================================================
# 段階 1: ヘッダセル分類
# =============================================================================


def detect_header_cells(ws: Any, header_row: int, wb_dataonly: Any) -> tuple[dict[str, Any], float, list[str]]:
    """Classify meta cells above the table header row.

    分類器で各セルを team_name / datetime / equipment / facility / theme に振り分ける。
    値と数式の両方を参照し、日付型セルも正しく検出する。
    """
    warnings: list[str] = []
    hits: dict[str, tuple[str, str, float]] = {}
    limit = max(header_row - 1, 1)
    ws_v = wb_dataonly[ws.title] if wb_dataonly else ws
    for row in ws.iter_rows(min_row=1, max_row=limit):
        for cell in row:
            raw = cell.value
            value_text = cell_text(raw)
            if not value_text:
                continue
            v_cell = ws_v.cell(row=cell.row, column=cell.column)
            display = cell_text(v_cell.value) if v_cell.value is not None else value_text
            coord = cell.coordinate

            if isinstance(v_cell.value, datetime) or DATETIME_RE.search(display):
                hits.setdefault("datetime", (coord, display, 0.95))
                continue
            if EQUIPMENT_PREFIX_RE.search(display):
                hits.setdefault("equipment", (coord, display, 0.95))
                continue
            if "equipment" not in hits and EQUIPMENT_ITEMS_RE.search(display) and len(display) > 10:
                hits.setdefault("equipment", (coord, display, 0.7))
                continue
            if FACILITY_RE.search(display) and not EQUIPMENT_PREFIX_RE.search(display):
                hits.setdefault("facility", (coord, display, 0.85))
                continue
            if THEME_RE.search(display) and not EQUIPMENT_PREFIX_RE.search(display):
                hits.setdefault("theme", (coord, display, 0.8))
                continue
            # team_name = 短いテキスト、まだ team_name 未確定
            if "team_name" not in hits and len(display) <= 15 and not DATETIME_RE.search(display):
                hits.setdefault("team_name", (coord, display, 0.6))

    header_cells: dict[str, Any] = {}
    for key in ("team_name", "datetime", "equipment", "facility", "theme"):
        if key in hits:
            coord, sample, conf = hits[key]
            spec: dict[str, Any] = {"cell": coord, "type": "text", "sample_value": sample, "confidence": conf}
            if key == "datetime":
                spec["type"] = "datetime_range"
            if key == "equipment":
                spec["prefix_regex"] = "^Equipment[:：]\\s*"
            header_cells[key] = spec
        else:
            warnings.append(f"header_cells.{key} not detected")

    conf_avg = sum(v[2] for v in hits.values()) / max(len(hits), 1) if hits else 0.0
    return header_cells, conf_avg, warnings


# =============================================================================
# 段階 2-3: table header 行 + 列マッピング
# =============================================================================


def detect_table_header_row(ws: Any) -> tuple[int | None, dict[str, str], float]:
    """Find the header row and map columns to canonical keys.

    Category/Times/Distance/... のキーワードが最も多く並ぶ行をヘッダとみなす。
    英日別名 (種目/回数/距離/セット/サイクル/内容/器具) にも対応。
    """
    best_row: int | None = None
    best_columns: dict[str, str] = {}
    best_score = 0
    for row in ws.iter_rows(min_row=1, max_row=min(30, ws.max_row)):
        found: dict[str, str] = {}
        for cell in row:
            key = TABLE_HEADER_ALIASES.get(normalize(cell.value))
            if key and key not in found:
                found[key] = cell.column_letter
        if len(found) > best_score:
            best_score = len(found)
            best_columns = found
            best_row = row[0].row
    if best_row is None or best_score < 3:
        return None, {}, 0.0
    conf = min(1.0, best_score / 5.0)
    return best_row, best_columns, conf


def enrich_columns_with_headers(ws: Any, header_row: int, base_columns: dict[str, str]) -> dict[str, dict[str, Any]]:
    """Attach header labels and infer types for each column.

    ヘッダラベル (I5=' ', J5='変換' のような空白/記号) は生値のまま保持し、
    Excel Table 構造化参照に必要な情報を揃える。
    """
    out: dict[str, dict[str, Any]] = {}
    for key, col_letter in base_columns.items():
        label = ws[f"{col_letter}{header_row}"].value
        spec: dict[str, Any] = {"col": col_letter, "header_label": "" if label is None else str(label)}
        if key in {"times", "distance", "set"}:
            spec["type"] = "int"
        elif key == "cycle":
            spec["type"] = "cycle_string"
        out[key] = spec
    return out


def detect_extra_formula_columns(ws: Any, header_row: int, mapped_letters: set[str], body_start: int, sample_rows: int = 8) -> dict[str, dict[str, Any]]:
    """Detect unmapped columns that carry formulas (subtotal / cycle_secs / block_secs / elapsed).

    ヘッダラベルが空白のみでも数式が並んでいれば数式列としてマップする。
    列名は数式パターンからさらに推測 (段階 6 で最終決定)。
    """
    extras: dict[str, dict[str, Any]] = {}
    max_col = ws.max_column or 12
    max_row = min(body_start + sample_rows, ws.max_row or body_start + sample_rows)
    for col_index in range(1, max_col + 1):
        letter = column_letter(col_index)
        if letter in mapped_letters:
            continue
        formula_hits = 0
        sample: str | None = None
        for r in range(body_start, max_row + 1):
            v = ws.cell(row=r, column=col_index).value
            if isinstance(v, str) and v.startswith("="):
                formula_hits += 1
                sample = sample or v
        if formula_hits >= 2:
            label_cell = ws.cell(row=header_row, column=col_index).value
            extras[f"_formula_col_{letter}"] = {
                "col": letter,
                "header_label": "" if label_cell is None else str(label_cell),
                "type": "formula",
                "sample_formula": sample,
            }
    return extras


# =============================================================================
# 段階 4: セクションヘッダパターン
# =============================================================================


def detect_section_header_rule(ws: Any, body_start: int, columns: dict[str, dict[str, Any]]) -> tuple[dict[str, Any], float, list[str]]:
    """Learn the section header row pattern (only one column has text).

    body 行を走査し、他列が空でひとつの列だけ文字列を持つ行のパターンを集計する。
    """
    warnings: list[str] = []
    letters = [spec["col"] for spec in columns.values()]
    lonely_col_counter: Counter[str] = Counter()
    section_examples: list[str] = []
    for r in range(body_start, min(body_start + 200, (ws.max_row or body_start) + 1)):
        non_empty_cols: list[str] = []
        text_only_col: str | None = None
        for letter in letters:
            v = ws[f"{letter}{r}"].value
            if v is None:
                continue
            if isinstance(v, str) and v.startswith("="):
                continue  # 数式は無視
            non_empty_cols.append(letter)
            if isinstance(v, str) and not v.startswith("="):
                text_only_col = letter
        if len(non_empty_cols) == 1 and text_only_col:
            v = ws[f"{text_only_col}{r}"].value
            if isinstance(v, str) and v.strip() and not TOTAL_MARKER_RE.match(v.strip()):
                lonely_col_counter[text_only_col] += 1
                if len(section_examples) < 8:
                    section_examples.append(v.strip())

    if not lonely_col_counter:
        warnings.append("section_header_rule: no section header rows detected")
        return {"detect": "only_column_B_has_text", "examples": []}, 0.3, warnings

    (target_col, count), *_ = lonely_col_counter.most_common(1)
    excluded = [letter for letter in letters if letter != target_col]
    conf = min(1.0, count / 4.0)
    return (
        {
            "detect": f"only_column_{target_col}_has_text",
            "target_column": target_col,
            "required_empty_columns": excluded,
            "clock_column_may_have_formula": True,
            "examples": section_examples,
        },
        conf,
        warnings,
    )


# =============================================================================
# 段階 5: 個別注記 bracket 検出
# =============================================================================


def detect_individual_note(ws: Any, body_start: int, description_col: str | None) -> tuple[dict[str, Any] | None, float, list[str]]:
    """Detect the individual-note bracket convention in the description column.

    Description 列内で `^【...】`, `^「...」`, `^\\[...\\]` などの頻度を数え、
    最頻の bracket を採用する。
    """
    warnings: list[str] = []
    if not description_col:
        warnings.append("individual_note: description column unknown")
        return None, 0.0, warnings

    counts: Counter[tuple[str, str, str]] = Counter()
    matched_rows = 0
    total_desc_rows = 0
    for r in range(body_start, min(body_start + 200, (ws.max_row or body_start) + 1)):
        v = ws[f"{description_col}{r}"].value
        if not isinstance(v, str) or not v.strip():
            continue
        total_desc_rows += 1
        for opener, closer, pattern in BRACKET_PAIRS:
            if re.match(pattern, v.strip()):
                counts[(opener, closer, pattern)] += 1
                matched_rows += 1
                break
    if not counts:
        warnings.append("individual_note: no bracket-prefixed rows found (may not use this convention)")
        return None, 0.4, warnings

    (opener, closer, pattern), count = counts.most_common(1)[0]
    conf = min(1.0, matched_rows / max(total_desc_rows, 1) * 3.0)
    return (
        {
            "column": description_col,
            "prefix_regex": pattern,
            "athlete_separator_regex": "[・、,／/]+",
            "opener": opener,
            "closer": closer,
        },
        conf,
        warnings,
    )


# =============================================================================
# 段階 6: 数式パターン照合
# =============================================================================


def classify_formula(formula: str, label_to_key: dict[str, str]) -> tuple[str | None, dict[str, Any]]:
    """Classify one formula string into a Descriptor v2 formula type.

    戻り値: ``(type, extra_spec)``. 判別不能なら ``(None, {})``.

    ``label_to_key`` はヘッダラベル (例 "Times") から canonical key (例 "times") への
    逆引き辞書。
    """
    if not formula or not formula.startswith("="):
        return None, {}
    text = formula
    refs = FORMULA_STRUCTURED_REF_RE.findall(text)
    ref_keys: list[str] = [label_to_key[label] for label in refs if label in label_to_key]

    # cycle_secs: TEXTJOIN + FIND("'") パターン
    if FORMULA_TEXTJOIN_RE.search(text) and "FIND(\"'\"" in text and "FIND(\"\"\"\"" in text:
        return "parse_time_string", {"input_format": "cycle"}

    # sum_minus_dynamic_ranges: SUM(table[...]) + "-SUM(" が続く
    if FORMULA_SUM_RE.search(text) and "-SUM(" in text.upper():
        return "sum_minus_dynamic_ranges", {"note": "SUM minus subtracted ranges detected"}

    # sum_column: 単一 SUM(table[...]) のみ
    if FORMULA_SUM_RE.search(text) and text.upper().count("SUM(") == 1:
        return "sum_column", {}

    # cumulative_from_column: =<letter>{n}+... パターン
    if re.match(r"^=\s*([A-Z]+)(\d+)\s*\+", text) and refs:
        return "cumulative_from_column", {}

    # product: * のみで結ばれた構造化参照
    core = text.lstrip("=")
    if refs and "*" in core and "+" not in core and "SUM" not in core.upper() and "/" not in core and "-" not in core:
        parts = core.split("*")
        operand_keys: list[str] = []
        for part in parts:
            m2 = FORMULA_STRUCTURED_REF_RE.search(part)
            if not m2:
                continue
            label = m2.group(1)
            if label in label_to_key:
                operand_keys.append(label_to_key[label])
        if len(operand_keys) >= 2:
            return "product", {"operands": operand_keys}

    return None, {}


def detect_formulas(ws: Any, body_start: int, header_row: int, columns: dict[str, dict[str, Any]], extras: dict[str, dict[str, Any]]) -> tuple[dict[str, Any], float, list[str]]:
    """Read formulas from a sample of body rows and classify per column.

    ラベル → canonical key の逆引き辞書を渡し、構造化参照を分類する。
    """
    warnings: list[str] = []
    # ラベル → canonical key の辞書 (columns_by_label)
    label_to_key: dict[str, str] = {}
    for key, spec in columns.items():
        label = spec.get("header_label", "") or spec["col"]
        label_to_key.setdefault(label, key)
    for key, spec in extras.items():
        label = spec.get("header_label", "") or spec["col"]
        label_to_key.setdefault(label, key)

    formula_specs: dict[str, dict[str, Any]] = {}
    max_row = min(body_start + 20, ws.max_row or body_start + 20)
    all_columns = {**columns, **extras}
    # col_letter → canonical name のライブ更新用マップ
    col_to_canonical: dict[str, str] = {}
    for key, spec in all_columns.items():
        col_letter = spec["col"]
        samples: list[str] = []
        for r in range(body_start, max_row + 1):
            v = ws[f"{col_letter}{r}"].value
            if isinstance(v, str) and v.startswith("="):
                samples.append(v)
        if not samples:
            continue
        classifications = [classify_formula(s, label_to_key) for s in samples]
        type_counter: Counter[str] = Counter(t for t, _ in classifications if t)
        if not type_counter:
            warnings.append(f"formulas.{key}: could not classify (col {col_letter})")
            continue
        best_type, _ = type_counter.most_common(1)[0]
        extra_spec = next((extra for t, extra in classifications if t == best_type), {})

        # canonical name を推測: elapsed / cycle_secs / block_secs / subtotal_m
        canonical_name = key
        if best_type == "cumulative_from_column":
            canonical_name = "elapsed"
            for f in samples:
                m = FORMULA_STRUCTURED_REF_RE.search(f)
                if m and m.group(1) in label_to_key:
                    add_key_label = label_to_key[m.group(1)]
                    extra_spec["add_column"] = col_to_canonical.get(all_columns.get(add_key_label, {}).get("col", ""), add_key_label)
                    break
            m2 = re.match(r"^=\s*([A-Z]+)(\d+)\s*\+", samples[0])
            if m2:
                extra_spec["seed_cell"] = f"{m2.group(1)}{max(1, body_start)}"
                extra_spec["seed_type"] = "time"
        elif best_type == "parse_time_string":
            canonical_name = "cycle_secs"
            for f in samples:
                m = FORMULA_STRUCTURED_REF_RE.search(f)
                if m and m.group(1) in label_to_key:
                    extra_spec["input"] = label_to_key[m.group(1)]
                    break
        elif best_type == "product":
            raw_operands = list(extra_spec.get("operands", []))
            # operands 内の一時名 (_formula_col_X) を canonical にリマップ
            remapped: list[str] = []
            for op in raw_operands:
                if op in col_to_canonical.values():
                    remapped.append(op)
                    continue
                spec_of_op = all_columns.get(op)
                if spec_of_op and spec_of_op["col"] in col_to_canonical:
                    remapped.append(col_to_canonical[spec_of_op["col"]])
                else:
                    remapped.append(op)
            extra_spec["operands"] = remapped
            if set(remapped) == {"times", "distance", "set"}:
                canonical_name = "subtotal_m"
            elif "cycle_secs" in remapped:
                canonical_name = "block_secs"

        formula_specs[canonical_name] = {
            "col": col_letter,
            "type": best_type,
            "on_blank_row": "continue_prev" if best_type == "cumulative_from_column" else "skip",
            "header_label": all_columns[key].get("header_label", "") or col_letter,
            **extra_spec,
        }
        col_to_canonical[col_letter] = canonical_name
        # label_to_key も canonical に上書き
        label_of_col = all_columns[key].get("header_label", "") or col_letter
        label_to_key[label_of_col] = canonical_name

    conf = min(1.0, len(formula_specs) / 4.0) if formula_specs else 0.0
    return formula_specs, conf, warnings


# =============================================================================
# 段階 7: TOTAL 行検出
# =============================================================================


def detect_total_row(ws: Any, body_start: int, all_columns: dict[str, dict[str, Any]]) -> tuple[dict[str, Any] | None, float, list[str]]:
    """Find the TOTAL row and classify its formula.

    "Total"/"合計"/"総距離" マーカーを含む列位置と、
    SUM 数式が置かれた列を組み合わせて記述する。
    ``all_columns`` は canonical + extras (数式のみ列) の合体辞書。
    """
    warnings: list[str] = []
    for r in range(body_start, (ws.max_row or body_start) + 1):
        marker_col = None
        marker_text = None
        for spec in all_columns.values():
            v = ws[f"{spec['col']}{r}"].value
            if isinstance(v, str) and TOTAL_MARKER_RE.match(v.strip()):
                marker_col = spec["col"]
                marker_text = v.strip()
                break
        if not marker_col:
            continue
        # 同じ行の SUM 数式列を探す
        for spec in all_columns.values():
            v = ws[f"{spec['col']}{r}"].value
            if isinstance(v, str) and v.startswith("=") and "SUM" in v.upper():
                type_ = "sum_minus_dynamic_ranges" if "-SUM(" in v.upper() else "sum_column"
                return (
                    {
                        "col": spec["col"],
                        "type": type_,
                        "of": "subtotal_m",
                        "subtract_filters": ["individual_note"] if type_ == "sum_minus_dynamic_ranges" else [],
                        "row_marker_column": marker_col,
                        "row_marker_text": marker_text,
                        "detected_row": r,
                    },
                    0.9,
                    warnings,
                )
        warnings.append(f"total row at row {r} has no SUM formula")
    return None, 0.0, warnings


# =============================================================================
# 段階 8: start_time_seed
# =============================================================================


def detect_start_time_seed(ws: Any, body_start: int, all_columns: dict[str, dict[str, Any]]) -> tuple[dict[str, Any] | None, float]:
    """Find a cell containing a time value near the top of the body.

    典型的には elapsed 列の先頭に開始時刻が入っている。
    canonical + extras (数式のみ列) の両方を走査する。
    """
    max_row = min(body_start + 3, ws.max_row or body_start + 3)
    for spec in all_columns.values():
        for r in range(body_start, max_row + 1):
            v = ws[f"{spec['col']}{r}"].value
            if isinstance(v, time):
                return (
                    {
                        "cell": f"{spec['col']}{r}",
                        "type": "time",
                        "example_value": v.strftime("%H:%M"),
                    },
                    0.9,
                )
    return None, 0.0


# =============================================================================
# 段階 9: descriptor 組立
# =============================================================================


def extract_sheet_styles(
    ws: Any,
    header_cells: dict[str, dict[str, Any]],
    header_row: int,
    body_start: int,
    columns: dict[str, dict[str, Any]],
    extras: dict[str, dict[str, Any]],
    formulas: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    """Capture per-cell styles + column widths from the reference sheet.

    書き出す情報:

    - header_cells の各名前付きセル (team_name / datetime / equipment / ...)
    - table header 行 (代表として最初の列のスタイルを採用)
    - body_row_default: 列ごとのスタイル (最初の非セクション・非空行から採取)
    - section_header_row: セクション行の代表スタイル
    - total_row: TOTAL 行の代表スタイル
    - column_widths / row_heights
    """
    styles: dict[str, Any] = {"header_cells": {}, "body_row_default": {"by_column": {}}}

    for name, spec in (header_cells or {}).items():
        cell = spec.get("cell")
        if cell:
            styles["header_cells"][name] = cell_style_to_dict(ws[cell])

    # table header 行 (最初の canonical 列で代表)
    first_col_letter = next(iter(columns.values()), {}).get("col") if columns else None
    if first_col_letter:
        styles["table_header_row"] = cell_style_to_dict(ws[f"{first_col_letter}{header_row}"])

    # 代表 body 行 (最初の "完全な" 行: description に文字列があり category も入っている)
    ref_row = None
    for r in range(body_start, min(body_start + 30, (ws.max_row or body_start) + 1)):
        first_val = ws[f"{first_col_letter}{r}"].value if first_col_letter else None
        desc_col = columns.get("description", {}).get("col")
        desc_val = ws[f"{desc_col}{r}"].value if desc_col else None
        if first_val and desc_val:
            ref_row = r
            break
    if ref_row:
        all_cols = {**columns, **extras}
        col_letter_to_key: dict[str, str] = {}
        # canonical / extras + formulas から key を確定
        for key, spec in all_cols.items():
            col_letter_to_key[spec["col"]] = key
        # formulas dict の canonical rename も反映
        for canonical, fspec in formulas.items():
            col_letter = fspec.get("col")
            if col_letter:
                col_letter_to_key[col_letter] = canonical
        for col_letter, key in col_letter_to_key.items():
            styles["body_row_default"]["by_column"][key] = cell_style_to_dict(ws[f"{col_letter}{ref_row}"])

    # section header 行 (最初のセクション: 1 列のみ text)
    for r in range(body_start, min(body_start + 30, (ws.max_row or body_start) + 1)):
        row_texts = 0
        section_col = None
        for spec in columns.values():
            v = ws[f"{spec['col']}{r}"].value
            if isinstance(v, str) and v.strip():
                row_texts += 1
                section_col = spec["col"]
        if row_texts == 1 and section_col:
            styles["section_header_row"] = cell_style_to_dict(ws[f"{section_col}{r}"])
            break

    # total 行 (formulas.total_marker で見つかれば)
    total_marker = formulas.get("total_marker") or {}
    total_col = total_marker.get("column")
    if total_col:
        for r in range(body_start, (ws.max_row or body_start) + 1):
            v = ws[f"{total_col}{r}"].value
            if isinstance(v, str) and v.strip().lower() in {"total", "合計", "計"}:
                styles["total_row"] = cell_style_to_dict(ws[f"{total_col}{r}"])
                break

    dims = sheet_dimensions_to_dict(ws)
    styles["column_widths"] = dims["column_widths"]
    styles["row_heights"] = dims["row_heights"]
    return styles


def extract_layout(source: "Path | bytes", sheet_name: str | None = None, layout_id: str | None = None, *, display_name: str | None = None, template_source_meta: dict[str, Any] | None = None) -> dict[str, Any]:
    """Extract a Descriptor v2 from one sheet of a workbook.

    ``source`` は Path (ローカル xlsx) か bytes (ネット経由で取得した xlsx バイト列) を受け付ける。
    ``sheet_name`` 未指定時は 最後のシート を使う (直近テンプレート想定)。
    ``display_name``/``template_source_meta`` は URL 由来の場合に上書き用。
    """
    import io as _io
    openpyxl = import_openpyxl()
    if isinstance(source, (bytes, bytearray)):
        wb = openpyxl.load_workbook(_io.BytesIO(bytes(source)), data_only=False)
        wb_v = openpyxl.load_workbook(_io.BytesIO(bytes(source)), data_only=True)
        source_stem = display_name or "sheets-source"
        source_display = display_name or "sheets-source"
    else:
        wb = openpyxl.load_workbook(source, data_only=False)
        wb_v = openpyxl.load_workbook(source, data_only=True)
        source_stem = Path(source).stem
        source_display = display_name or Path(source).name
    target = sheet_name or wb.sheetnames[-1]
    ws = wb[target]

    warnings: list[str] = []
    confidences: list[float] = []

    header_row, base_columns, header_conf = detect_table_header_row(ws)
    confidences.append(header_conf)
    if header_row is None:
        return {
            "$schema_version": "excel-layout-descriptor/v2",
            "layout_id": layout_id or source_stem,
            "detected_from": f"{source_display}#{target}",
            "confidence": 0.0,
            "warnings": ["table header row not detected"],
            "template_source": template_source_meta or {
                "xlsx_path": str(source).replace("\\", "/") if not isinstance(source, (bytes, bytearray)) else None,
                "sheet_name": target,
            },
        }
    body_start = header_row + 1

    columns = enrich_columns_with_headers(ws, header_row, base_columns)
    mapped_letters = {spec["col"] for spec in columns.values()}
    extras = detect_extra_formula_columns(ws, header_row, mapped_letters, body_start)

    header_cells, hc_conf, hc_warnings = detect_header_cells(ws, header_row, wb_v)
    confidences.append(hc_conf); warnings.extend(hc_warnings)

    section_rule, sec_conf, sec_warnings = detect_section_header_rule(ws, body_start, columns)
    confidences.append(sec_conf); warnings.extend(sec_warnings)

    description_col = columns.get("description", {}).get("col")
    ind_note, ind_conf, ind_warnings = detect_individual_note(ws, body_start, description_col)
    confidences.append(ind_conf); warnings.extend(ind_warnings)

    formulas, fmla_conf, fmla_warnings = detect_formulas(ws, body_start, header_row, columns, extras)
    confidences.append(fmla_conf); warnings.extend(fmla_warnings)

    # canonical + extras の統合ビュー (数式のみ列も含む TOTAL / seed 検出用)
    all_columns_view: dict[str, dict[str, Any]] = {**columns, **extras}

    total_spec, total_conf, total_warnings = detect_total_row(ws, body_start, all_columns_view)
    confidences.append(total_conf); warnings.extend(total_warnings)
    total_row_number = None
    if total_spec:
        total_row_number = total_spec.pop("detected_row", None)
        formulas["total_marker"] = {
            "column": total_spec.pop("row_marker_column", None),
            "text": total_spec.pop("row_marker_text", "Total"),
            "template_row": total_row_number,
        }
        formulas["total_distance"] = total_spec

    seed_spec, seed_conf = detect_start_time_seed(ws, body_start, all_columns_view)
    confidences.append(seed_conf)

    styles = extract_sheet_styles(ws, header_cells, header_row, body_start, columns, extras, formulas)

    # merged columns (extras を canonical に統合したので不要)
    canonical_columns = columns

    overall_conf = sum(confidences) / max(len(confidences), 1)

    descriptor: dict[str, Any] = {
        "$schema_version": "excel-layout-descriptor/v2",
        "layout_id": layout_id or source_stem,
        "display_name": f"Auto-detected from {source_display}",
        "detected_from": f"{source_display}#{target}",
        "sample_sheets": [target],
        "confidence": round(overall_conf, 3),
        "confidence_breakdown": {
            "table_header": round(header_conf, 3),
            "header_cells": round(hc_conf, 3),
            "section_header_rule": round(sec_conf, 3),
            "individual_note": round(ind_conf, 3),
            "formulas": round(fmla_conf, 3),
            "total_row": round(total_conf, 3),
            "start_time_seed": round(seed_conf, 3),
        },
        "warnings": warnings,
        "header_cells": header_cells,
        "table": {
            "header_row": header_row,
            "body_start_row": body_start,
            "columns": canonical_columns,
            "table_name_prefix": "Menu",
            "table_style": "TableStyleMedium2",
        },
        "section_header_rule": section_rule,
        "individual_note": ind_note or {"column": description_col, "prefix_regex": r"^【([^】]+)】", "athlete_separator_regex": "[・、,／/]+"},
        "empty_row_as_separator": True,
        "formulas": formulas,
        "start_time_seed": seed_spec or {"cell": None, "type": "time"},
        "pace_table": None,
        "notes_block": {
            "style": "inline",
            "location": "description_column",
            "syntax": "個別注記 prefix_regex",
            "separate_pace_table": False,
        },
        "styles": styles,
        "template_source": template_source_meta or {
            "xlsx_path": str(source).replace("\\", "/") if not isinstance(source, (bytes, bytearray)) else None,
            "sheet_name": target,
            "_note": "Writer が指定されればこのシートを複製 → body クリア → 書き込み で見た目 100% 再現。パスが解決できない場合は styles 辞書でフォールバック。",
        },
    }
    # ensure sheet_name always present in template_source
    if descriptor["template_source"].get("sheet_name") is None:
        descriptor["template_source"]["sheet_name"] = target
    return descriptor


GOOGLE_SHEETS_URL_RE = re.compile(r"docs\.google\.com/spreadsheets/d/(?P<id>[a-zA-Z0-9_-]+)")


class GoogleSheetsAccessError(RuntimeError):
    """Google Sheets の共有設定 or URL 不備で xlsx が取得できないときの例外。

    メッセージには対処手順を含める (「リンクを知っている全員(閲覧可)」に変更, etc.)。
    """


def parse_google_sheets_url(url: str) -> tuple[str, str | None]:
    """URL から (doc_id, gid) を抽出する。gid は URL に含まれなければ None。"""
    m = GOOGLE_SHEETS_URL_RE.search(url)
    if not m:
        raise ValueError(f"not a Google Sheets URL: {url}")
    doc_id = m.group("id")
    gid_m = re.search(r"[#&?]gid=(\d+)", url)
    gid = gid_m.group(1) if gid_m else None
    return doc_id, gid


def fetch_google_sheet_xlsx_bytes(url: str, gid: str | None = None, timeout: int = 30) -> bytes:
    """Public Google Sheets を xlsx バイト列として取得する (ディスクに書かない)。

    URL は共有設定が「リンクを知っている全員 (閲覧可)」以上である必要がある。
    非公開シートの場合は Google がログイン HTML を返してくるので、それを検知して
    ``GoogleSheetsAccessError`` を投げ、ユーザに共有設定変更を促す。
    """
    import urllib.request
    import urllib.error

    doc_id, url_gid = parse_google_sheets_url(url)
    gid = gid or url_gid
    export_url = f"https://docs.google.com/spreadsheets/d/{doc_id}/export?format=xlsx"
    if gid:
        export_url += f"&gid={gid}"
    req = urllib.request.Request(export_url, headers={"User-Agent": "spagent/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            data = resp.read()
            content_type = resp.headers.get("Content-Type", "")
    except urllib.error.HTTPError as exc:
        if exc.code in (401, 403):
            raise GoogleSheetsAccessError(
                f"Sheets アクセス拒否 (HTTP {exc.code})。共有設定を「リンクを知っている全員 (閲覧可)」以上にしてください。URL: {url}"
            ) from exc
        if exc.code == 404:
            raise GoogleSheetsAccessError(f"Sheets が見つかりません (HTTP 404): {url}") from exc
        raise
    # 成功時でも Google はログインページ HTML を返してくることがある (非公開シート)
    if data.startswith(b"PK"):
        return data
    if b"<html" in data[:2048].lower() or "text/html" in content_type.lower():
        raise GoogleSheetsAccessError(
            "Google がログイン HTML を返しました。シートが非公開の可能性が高いです。\n"
            "対処: 共有設定を「リンクを知っている全員 (閲覧可)」以上に変更するか、\n"
            "      手動で ファイル → ダウンロード → Microsoft Excel (.xlsx) で保存してローカルパス指定してください。\n"
            f"URL: {url}"
        )
    raise GoogleSheetsAccessError(f"downloaded content is not a valid xlsx (first 8 bytes: {data[:8]!r})")


def list_google_sheet_tabs(xlsx_bytes: bytes) -> list[str]:
    """バイト列 (xlsx) からシート名リストを返す。タブ選択 UI 用。"""
    import io as _io
    openpyxl = import_openpyxl()
    wb = openpyxl.load_workbook(_io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def build_parser() -> argparse.ArgumentParser:
    """Build the CLI parser.

    xlsx パス (または Google Sheets URL) 、シート指定、出力先、layout_id を受け取る。
    """
    parser = argparse.ArgumentParser(description="Auto-detect an Excel Layout Descriptor v2 from an xlsx file or a public Google Sheets URL.")
    parser.add_argument("source", help="Coach's xlsx path OR a public Google Sheets URL.")
    parser.add_argument("--sheet-name", help="Target sheet name (defaults to last sheet).")
    parser.add_argument("--sheet-gid", help="Google Sheets tab gid (used only when source is a Sheets URL).")
    parser.add_argument("--layout-id", help="layout_id for the produced descriptor.")
    parser.add_argument("--out", type=Path, help="Output descriptor JSON path.")
    parser.add_argument("--cache-to", type=Path, default=None, help="Optional path to persist downloaded xlsx bytes (Sheets URL のみ; デフォルトは in-memory のみ).")
    parser.add_argument("--list-sheets", action="store_true", help="Sheets URL からシート名一覧を表示して終了 (どのタブを --sheet-name で指定するか確認する用).")
    return parser


def main(argv: list[str] | None = None) -> int:
    """CLI entrypoint.

    成功時は descriptor を stdout か --out に書き出す。
    xlsx path も Google Sheets URL も受け付ける。Sheets の場合はメモリ内で処理し、
    ``--cache-to`` 指定時のみディスクに保存する。
    """
    args = build_parser().parse_args(argv)
    source = args.source
    is_url = isinstance(source, str) and source.lower().startswith(("http://", "https://"))
    try:
        if is_url:
            if not GOOGLE_SHEETS_URL_RE.search(source):
                raise ValueError("only Google Sheets URLs are supported for remote source")
            xlsx_bytes = fetch_google_sheet_xlsx_bytes(source, gid=args.sheet_gid)
            if args.cache_to:
                args.cache_to.parent.mkdir(parents=True, exist_ok=True)
                args.cache_to.write_bytes(xlsx_bytes)
                print(f"cached: {args.cache_to}", file=sys.stderr)
            if args.list_sheets:
                for name in list_google_sheet_tabs(xlsx_bytes):
                    print(name)
                return 0
            doc_id, _ = parse_google_sheets_url(source)
            default_layout_id = args.layout_id or f"sheets-{doc_id[:8]}"
            template_meta = {
                "xlsx_path": str(args.cache_to).replace("\\", "/") if args.cache_to else None,
                "sheet_name": None,  # extract_layout が埋める
                "source_url": source,
                "source_kind": "google_sheets",
                "_note": "オンライン Sheets 由来。Writer 側で xlsx_path が None のときは styles/DescriptorV2 経由でフォールバック。",
            }
            descriptor = extract_layout(
                xlsx_bytes,
                sheet_name=args.sheet_name,
                layout_id=default_layout_id,
                display_name=f"sheets:{doc_id[:8]}",
                template_source_meta=template_meta,
            )
        else:
            xlsx_path = Path(source)
            descriptor = extract_layout(xlsx_path, args.sheet_name, args.layout_id)
    except GoogleSheetsAccessError as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 2
    except Exception as exc:  # noqa: BLE001 - CLI boundary
        print(f"error: {exc}", file=sys.stderr)
        return 1
    payload = json.dumps(descriptor, ensure_ascii=False, indent=2)
    if args.out:
        args.out.parent.mkdir(parents=True, exist_ok=True)
        args.out.write_text(payload + "\n", encoding="utf-8")
        print(args.out)
    else:
        print(payload)
    return 0


if __name__ == "__main__":
    sys.exit(main())
