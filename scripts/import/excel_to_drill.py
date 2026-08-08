"""Convert Excel drill sheets to JSON records.

ドリル名、目的、説明、フォーカスポイント、よくあるミスを Excel から抽出する。
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any


FIELD_ALIASES = {
    "name": "name",
    "drill": "name",
    "purpose": "purpose",
    "description": "description",
    "focuspoints": "focus_points",
    "focus": "focus_points",
    "commonmistakes": "common_mistakes",
    "mistakes": "common_mistakes",
}


def import_openpyxl() -> Any:
    """Import openpyxl with a helpful error.

    依存がない場合はインストール方法を含む例外を出す。
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required. Install with: python -m pip install openpyxl") from exc
    return openpyxl


def normalize(value: Any) -> str:
    """Normalize text for field matching.

    空白や記号を除いて小文字化する。
    """
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def text(value: Any) -> str:
    """Convert a cell value to text.

    None は空文字として扱う。
    """
    return str(value).strip() if value is not None else ""


def find_header_row(ws: Any) -> int | None:
    """Find a drill table header row.

    name/purpose/description 等の見出しが複数ある行を探す。
    """
    for row in ws.iter_rows():
        hits = sum(1 for cell in row if normalize(cell.value) in FIELD_ALIASES)
        if hits >= 2:
            return row[0].row
    return None


def parse_focus(value: str) -> list[str]:
    """Split focus or mistake text into a list.

    改行、セミコロン、読点で簡易分割する。
    """
    return [part.strip() for part in re.split(r"[\n;；、]+", value) if part.strip()]


def parse_sheet(ws: Any, source_path: Path) -> list[dict[str, Any]]:
    """Parse one worksheet into drill records.

    表形式がない場合はシート全体を 1 件のドリル候補として扱う。
    """
    header_row = find_header_row(ws)
    drills: list[dict[str, Any]] = []
    if header_row is None:
        values = [text(cell.value) for row in ws.iter_rows() for cell in row if text(cell.value)]
        if values:
            drills.append(
                {
                    "id": f"{source_path.stem}-{ws.title}",
                    "name": values[0],
                    "purpose": "",
                    "description": "\n".join(values[1:]),
                    "focus_points": [],
                    "common_mistakes": [],
                    "source_type": "excel",
                    "source_path": str(source_path),
                    "sheet_name": ws.title,
                }
            )
        return drills

    headers = [FIELD_ALIASES.get(normalize(cell.value), normalize(cell.value) or f"col_{cell.column}") for cell in ws[header_row]]
    for row in ws.iter_rows(min_row=header_row + 1):
        raw_values = [text(cell.value) for cell in row]
        if not any(raw_values):
            continue
        item = {headers[index]: value for index, value in enumerate(raw_values) if index < len(headers)}
        name = item.get("name", "")
        if not name:
            continue
        drills.append(
            {
                "id": f"{source_path.stem}-{len(drills) + 1}",
                "name": name,
                "purpose": item.get("purpose", ""),
                "description": item.get("description", ""),
                "focus_points": parse_focus(item.get("focus_points", "")),
                "common_mistakes": parse_focus(item.get("common_mistakes", "")),
                "source_type": "excel",
                "source_path": str(source_path),
                "sheet_name": ws.title,
            }
        )
    return drills


def convert_workbook(path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    """Convert workbook sheets to drill records.

    指定がなければ全シートを処理する。
    """
    openpyxl = import_openpyxl()
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=False)
    worksheets = [workbook[sheet_name]] if sheet_name else list(workbook.worksheets)
    records: list[dict[str, Any]] = []
    for ws in worksheets:
        records.extend(parse_sheet(ws, path))
    return records


def write_output(records: list[dict[str, Any]], out_path: Path | None) -> None:
    """Write JSON output.

    `--out` がある場合はファイルへ保存する。
    """
    payload = json.dumps(records, ensure_ascii=False, indent=2)
    if out_path:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(payload + "\n", encoding="utf-8")
    else:
        print(payload)


def build_parser() -> argparse.ArgumentParser:
    """Build the command-line parser.

    Excel ファイルと任意の出力先・シート名を受け取る。
    """
    parser = argparse.ArgumentParser(description="Convert an Excel workbook to drill JSON.")
    parser.add_argument("path", type=Path, help="Input .xlsx path.")
    parser.add_argument("--out", type=Path, help="Output JSON file.")
    parser.add_argument("--sheet-name", help="Only convert one sheet.")
    return parser


def main(argv: list[str] | None = None) -> int:
    """Run the Excel drill converter.

    失敗時は stderr に理由を表示する。
    """
    args = build_parser().parse_args(argv)
    try:
        records = convert_workbook(args.path, args.sheet_name)
        write_output(records, args.out)
    except Exception as exc:  # noqa: BLE001 - CLI boundary
        print(f"error: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
