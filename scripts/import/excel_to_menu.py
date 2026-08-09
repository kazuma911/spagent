"""Convert Excel swim menus to JSON records.

Excel の title/date/facility/theme/header/body/TOTAL という慣習を読み取り、
メニュー JSON の配列として出力する。
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable


HEADER_ALIASES = {
    "category": "category",
    "times": "times",
    "distance": "distance",
    "set": "set",
    "cycle": "cycle",
    "description": "description",
    "gears": "gears",
    "subtotal": "subtotal",
}


def import_openpyxl() -> Any:
    """Import openpyxl with a helpful error.

    依存がない場合はインストール方法を含む例外にする。
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required. Install with: python -m pip install openpyxl") from exc
    return openpyxl


def normalize(value: Any) -> str:
    """Normalize a cell value for matching.

    空白や記号を削って小文字にする。
    """
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def text(value: Any) -> str:
    """Convert a cell value to display text.

    日付型は ISO 形式に変換する。
    """
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip() if value is not None else ""


def row_values(row: Iterable[Any]) -> list[Any]:
    """Return raw values from a worksheet row.

    openpyxl の cell オブジェクトから値だけを取り出す。
    """
    return [cell.value for cell in row]


def find_header_row(ws: Any) -> int | None:
    """Find the likely menu header row.

    Category/Times/Distance などの見出しが複数ある行を探す。
    """
    for row in ws.iter_rows():
        values = row_values(row)
        hits = sum(1 for value in values if normalize(value) in HEADER_ALIASES)
        if hits >= 3:
            return row[0].row
    return None


def _first_row_text(row_texts: list[str]) -> str:
    """Return a compact string of non-empty cell texts in one row.

    行内の非空セルを空白 1 つで連結し、レイアウト非依存の判定用文字列にする。
    """
    return " ".join(value for value in row_texts if value).strip()


_DATE_RE = re.compile(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})")
_FACILITY_HINT_RE = re.compile(
    r"(SCM|LCM|\d{2,3}\s*m|pool|プール|センター|アクアティクス|温水|ジム|体育館|施設|場所|facility)",
    re.IGNORECASE,
)
_THEME_HINT_RE = re.compile(
    r"(Phase\s*[A-Z]|D-\s*\d+|Taper|Threshold|Sprint|Recovery|Base|USRPT|Broken|Descending"
    r"|Aerobic|Race\s*Pace|Endurance|IM\s*day|Kick\s*day|Drill\s*day|VO2|EN[123]|SP[12]"
    r"|大会|試合|レース|基礎期|準備期|移行期|テーマ|theme)",
    re.IGNORECASE,
)


def extract_meta(ws: Any, header_row: int | None) -> dict[str, str]:
    """Extract title, date, facility, theme, team_name, and equipment.

    ヘッダ行の直前までの任意レイアウト（列 A 起点でも B 起点でも可）に対応する。
    キーワード辞書だけでは分類できないため、正規表現ヒントで判別する。
    """
    limit = (header_row or 6) - 1
    rows = [[text(value) for value in row_values(row)] for row in ws.iter_rows(min_row=1, max_row=max(limit, 1))]
    row_texts = [_first_row_text(row) for row in rows]

    meta: dict[str, str] = {}

    for row_text in row_texts:
        if not row_text:
            continue
        # 明示ラベル `Equipment: ...` は最優先で拾う
        equipment_match = re.search(r"equipment\s*[:：]\s*(.+)", row_text, re.IGNORECASE)
        if equipment_match and "equipment" not in meta:
            meta["equipment"] = equipment_match.group(1).strip()

        # 日付を確実に抽出 (YYYY-MM-DD へ正規化、括弧混入を防ぐ)
        if "date" not in meta:
            date_match = _DATE_RE.search(row_text)
            if date_match:
                y, m, d = date_match.groups()
                meta["date"] = f"{int(y):04d}-{int(m):02d}-{int(d):02d}"

        # facility 判定 (プール施設らしいキーワードを含む)
        if "facility" not in meta and _FACILITY_HINT_RE.search(row_text):
            # equipment 行を facility として拾わない
            if not re.search(r"equipment\s*[:：]", row_text, re.IGNORECASE):
                meta["facility"] = row_text

        # theme 判定 (フェーズ / メソッド / 大会キーワードを含む)
        if "theme" not in meta and _THEME_HINT_RE.search(row_text):
            if not re.search(r"equipment\s*[:：]", row_text, re.IGNORECASE):
                meta["theme"] = row_text

    # team_name: 最初の短い (<= 12 文字) 行で date/facility/theme に採用されていないもの
    used = {meta.get("date", ""), meta.get("facility", ""), meta.get("theme", ""), meta.get("equipment", "")}
    for row_text in row_texts:
        if not row_text or row_text in used:
            continue
        if len(row_text) <= 12 and not _DATE_RE.search(row_text):
            meta.setdefault("team_name", row_text)
            break

    # title: 短い team_name とは別に、theme か team_name か sheet 名で埋める
    meta.setdefault("title", meta.get("theme") or meta.get("team_name") or ws.title)

    return meta


def safe_int(value: Any) -> int:
    """Extract an integer from a cell value.

    距離計算用に数値部分だけを取得する。
    """
    if isinstance(value, (int, float)):
        return int(value)
    match = re.search(r"\d+", str(value or ""))
    return int(match.group(0)) if match else 0


def parse_sheet(ws: Any, source_path: Path) -> dict[str, Any]:
    """Parse one worksheet as a menu record.

    TOTAL 行に到達したら本文の読み取りを終了する。
    """
    header_row = find_header_row(ws)
    meta = extract_meta(ws, header_row)
    structure: list[dict[str, Any]] = []
    if header_row is None:
        return {
            "id": f"{source_path.stem}-{ws.title}",
            "title": meta.get("title") or ws.title,
            "date": meta.get("date"),
            "facility": meta.get("facility"),
            "theme": meta.get("theme"),
            "total_distance": 0,
            "structure": structure,
            "source_type": "excel",
            "source_path": str(source_path),
            "sheet_name": ws.title,
            "warnings": ["header row was not detected"],
        }

    headers = [HEADER_ALIASES.get(normalize(cell.value), normalize(cell.value) or f"col_{cell.column}") for cell in ws[header_row]]
    for row in ws.iter_rows(min_row=header_row + 1):
        values = row_values(row)
        if not any(text(value) for value in values):
            continue
        if any(str(value).strip().upper() == "TOTAL" for value in values if value is not None):
            break
        item = {headers[index]: text(value) for index, value in enumerate(values) if index < len(headers)}
        if any(item.values()):
            item["set_no"] = str(len(structure) + 1)
            subtotal = safe_int(item.get("subtotal"))
            if not subtotal:
                subtotal = (safe_int(item.get("times")) or 1) * safe_int(item.get("distance"))
            item["estimated_distance"] = subtotal
            structure.append(item)
    total_distance = sum(safe_int(item.get("estimated_distance")) for item in structure)
    sheet_name = ws.title.strip()
    return {
        "id": f"{meta.get('date') or source_path.stem}-{sheet_name}",
        "title": meta.get("title") or sheet_name,
        "date": meta.get("date"),
        "team_name": meta.get("team_name"),
        "facility": meta.get("facility"),
        "theme": meta.get("theme"),
        "equipment": meta.get("equipment"),
        "total_distance": total_distance,
        "structure": structure,
        "source_type": "excel",
        "source_path": str(source_path),
        "sheet_name": sheet_name,
    }


def convert_workbook(path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    """Convert workbook sheets to menu records.

    `--sheet-name` が指定された場合はそのシートだけを処理する。
    """
    openpyxl = import_openpyxl()
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=False)
    worksheets = [workbook[sheet_name]] if sheet_name else list(workbook.worksheets)
    return [parse_sheet(ws, path) for ws in worksheets]


def write_output(records: list[dict[str, Any]], out_path: Path | None) -> None:
    """Write JSON output to file or stdout.

    `--out` があればファイル、なければ標準出力に出す。
    """
    payload = json.dumps(records, ensure_ascii=False, indent=2)
    if out_path:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(payload + "\n", encoding="utf-8")
    else:
        print(payload)


def build_parser() -> argparse.ArgumentParser:
    """Build the command-line parser.

    Excel パス、出力先、対象シートを受け取る。
    """
    parser = argparse.ArgumentParser(description="Convert an Excel workbook to menu JSON.")
    parser.add_argument("path", type=Path, help="Input .xlsx path.")
    parser.add_argument("--out", type=Path, help="Output JSON file.")
    parser.add_argument("--sheet-name", help="Only convert one sheet.")
    return parser


def main(argv: list[str] | None = None) -> int:
    """Run the Excel menu converter.

    変換に失敗した場合は非ゼロ終了する。
    """
    args = build_parser().parse_args(argv)
    try:
        records = convert_workbook(args.path, args.sheet_name)
        write_output(records, args.out)
    except Exception as exc:  # noqa: BLE001 - CLI boundary
        print(f"error: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
