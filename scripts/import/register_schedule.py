"""Register competitions and training sessions into spagent.

Three modes are supported so any coach can register their schedule:

* ``--mode manual`` — interactive dialog (stdin prompts).
* ``--mode file`` — extract raw candidates from an Excel/PDF placed in
  ``data/inbox/schedule/`` (the LLM in a Workflow F conversation reviews
  the candidates JSON and calls this script again with ``--merge``).
* ``--mode url``  — fetch a URL (HTML or PDF), same candidates flow.

Merging is idempotent:

* ``data/competitions.json`` — competitions upserted by ``id``.
* ``data/training-schedule.json`` — sessions upserted by ``date`` (only
  overwrites when the incoming entry provides a new ``block`` or
  ``focus``; existing per-athlete details are preserved).

Design constraint: this script does NOT call an LLM directly. The
LLM lives in the SKILL conversation. Scripts extract mechanical
content and dump candidates; the Workflow uses those in-context.

Usage examples::

    python scripts/import/register_schedule.py --mode manual
    python scripts/import/register_schedule.py --mode file --input data/inbox/schedule/foo.xlsx
    python scripts/import/register_schedule.py --mode url  --input https://example.org/comps.html
    python scripts/import/register_schedule.py --merge candidates.json --dry-run
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import urllib.error
import urllib.request
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable


def repo_root() -> Path:
    return Path(__file__).resolve().parents[2]


# ---------------------------------------------------------------------------
# Schema helpers
# ---------------------------------------------------------------------------

VALID_COURSES = {"SCM", "LCM", "TBD"}
VALID_BLOCKS = {"Acc", "Acc-peak", "Trans", "Real", "Taper", "★RACE", "external", "recovery", "unknown"}
DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def parse_date(text: str) -> date:
    return datetime.strptime(text, "%Y-%m-%d").date()


def load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def dump_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


# ---------------------------------------------------------------------------
# Interactive (manual) mode
# ---------------------------------------------------------------------------

def _ask(prompt: str, default: str | None = None, choices: Iterable[str] | None = None) -> str:
    while True:
        suffix = f" [{default}]" if default else ""
        val = input(f"  {prompt}{suffix}: ").strip()
        if not val and default is not None:
            val = default
        if choices and val not in choices:
            print(f"    (must be one of {sorted(choices)})")
            continue
        if val:
            return val


def _ask_optional(prompt: str) -> str | None:
    val = input(f"  {prompt} (enter to skip): ").strip()
    return val or None


def collect_manual_competition() -> dict[str, Any]:
    print("\n[新しい大会を登録] (Ctrl+C で中断)")
    id_ = _ask("id (英数ハイフン)", default=None)
    name = _ask("大会名", default=None)
    start = _ask("start_date (YYYY-MM-DD)")
    end = input(f"    end_date (YYYY-MM-DD) [{start}]: ").strip() or start
    course = _ask("course", default="LCM", choices=VALID_COURSES - {"TBD"})
    priority = _ask("priority (A/B/C)", default="B", choices={"A", "B", "C"})
    venue = _ask_optional("venue")
    city = _ask_optional("location_city")
    entries: list[dict[str, Any]] = []
    print("  entries (最低 1 件): 選手の種目を追加。空 Enter で終了。")
    while True:
        athlete = input("    athlete_id (空 Enter で終了): ").strip()
        if not athlete:
            break
        event = _ask("    event (例 100Fr, 200Ba)")
        target = _ask_optional("    target_time (mm:ss.xx or ss.xx)")
        entries.append({
            "athlete_id": athlete,
            "event": event,
            "target_time": target,
        })
    return {
        "id": id_,
        "name": name,
        "start_date": start,
        "end_date": end,
        "course": course,
        "priority": priority,
        "venue": venue,
        "location_city": city,
        "status": "planned",
        "entries": entries,
    }


def collect_manual_session() -> dict[str, Any]:
    print("\n[新しい練習セッションを登録] (Ctrl+C で中断)")
    d = _ask("date (YYYY-MM-DD)")
    dow_map = {0: "月", 1: "火", 2: "水", 3: "木", 4: "金", 5: "土", 6: "日"}
    dow_default = dow_map[parse_date(d).weekday()]
    dow = _ask("dow", default=dow_default)
    course = _ask("course", default="SCM", choices=VALID_COURSES)
    block = _ask("block (Acc/Trans/Real/Taper/★RACE 等)", default="Acc", choices=VALID_BLOCKS)
    focus = _ask("focus (今日の狙い)")
    managed_by = _ask("managed_by", default="self")
    return {
        "date": d, "dow": dow, "course": course,
        "block": block, "focus": focus, "managed_by": managed_by,
    }


def run_manual(root: Path, target: str) -> dict[str, Any]:
    candidates: dict[str, list[dict[str, Any]]] = {"competitions": [], "sessions": []}
    if target in ("competitions", "both"):
        while True:
            candidates["competitions"].append(collect_manual_competition())
            more = input("\n続けて大会を追加しますか？ [y/N]: ").strip().lower()
            if more != "y":
                break
    if target in ("training-schedule", "both"):
        while True:
            candidates["sessions"].append(collect_manual_session())
            more = input("\n続けて練習を追加しますか？ [y/N]: ").strip().lower()
            if more != "y":
                break
    return candidates


# ---------------------------------------------------------------------------
# File extraction mode (Excel / PDF)
# ---------------------------------------------------------------------------

def extract_from_excel(path: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel extraction: pip install openpyxl")
    wb = load_workbook(path, data_only=True, read_only=True)
    tables: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(cell not in (None, "") for cell in row):
                rows.append([str(cell) if cell is not None else "" for cell in row])
        if rows:
            tables.append({"sheet_name": ws.title, "rows": rows[:200]})
    return {"source": "excel", "path": str(path), "tables": tables}


def extract_from_pdf(path: Path) -> dict[str, Any]:
    try:
        import pdfplumber
    except ImportError:
        raise RuntimeError("pdfplumber is required for PDF extraction: pip install pdfplumber")
    pages: list[dict[str, Any]] = []
    with pdfplumber.open(path) as pdf:
        for i, page in enumerate(pdf.pages, 1):
            text = page.extract_text() or ""
            tables = [t for t in (page.extract_tables() or []) if t]
            pages.append({"page": i, "text": text[:5000], "tables": tables})
    return {"source": "pdf", "path": str(path), "pages": pages}


def run_file(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"file not found: {path}")
    ext = path.suffix.lower()
    if ext in (".xlsx", ".xlsm", ".xls"):
        return extract_from_excel(path)
    if ext == ".pdf":
        return extract_from_pdf(path)
    if ext in (".csv", ".tsv", ".txt", ".md"):
        return {"source": ext.lstrip("."), "path": str(path), "text": path.read_text(encoding="utf-8", errors="replace")[:20000]}
    raise ValueError(f"unsupported file type: {ext} (supported: .xlsx .pdf .csv .tsv .txt .md)")


# ---------------------------------------------------------------------------
# URL fetch mode
# ---------------------------------------------------------------------------

def run_url(url: str) -> dict[str, Any]:
    if not url.startswith(("http://", "https://")):
        raise ValueError(f"URL must start with http:// or https://: {url}")
    req = urllib.request.Request(url, headers={"User-Agent": "spagent-schedule-fetch/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            content_type = resp.headers.get_content_type() or ""
            raw = resp.read()
    except urllib.error.URLError as exc:
        raise RuntimeError(f"fetch failed: {exc}") from exc
    if "pdf" in content_type or url.lower().endswith(".pdf"):
        tmp = repo_root() / "data" / "inbox" / "schedule" / "_fetched.pdf"
        tmp.parent.mkdir(parents=True, exist_ok=True)
        tmp.write_bytes(raw)
        result = extract_from_pdf(tmp)
        result["source_url"] = url
        return result
    text = raw.decode("utf-8", errors="replace")
    text_stripped = re.sub(r"<script[^>]*>.*?</script>", "", text, flags=re.S | re.I)
    text_stripped = re.sub(r"<style[^>]*>.*?</style>", "", text_stripped, flags=re.S | re.I)
    text_stripped = re.sub(r"<[^>]+>", " ", text_stripped)
    text_stripped = re.sub(r"\s+", " ", text_stripped).strip()
    return {"source": "url", "url": url, "content_type": content_type, "text": text_stripped[:20000]}


# ---------------------------------------------------------------------------
# Merge into destination JSON
# ---------------------------------------------------------------------------

def _validate_competition(c: dict[str, Any]) -> list[str]:
    errs: list[str] = []
    for f in ("id", "name", "start_date", "course"):
        if not c.get(f):
            errs.append(f"competition missing '{f}'")
    if c.get("start_date") and not DATE_RE.match(str(c["start_date"])):
        errs.append(f"start_date must be YYYY-MM-DD: {c['start_date']}")
    if c.get("course") and c["course"] not in VALID_COURSES:
        errs.append(f"course must be one of {VALID_COURSES}: {c['course']}")
    return errs


def _validate_session(s: dict[str, Any]) -> list[str]:
    errs: list[str] = []
    if not s.get("date") or not DATE_RE.match(str(s["date"])):
        errs.append(f"session date invalid: {s.get('date')}")
    if s.get("course") and s["course"] not in VALID_COURSES:
        errs.append(f"session course must be one of {VALID_COURSES}: {s['course']}")
    if s.get("block") and s["block"] not in VALID_BLOCKS:
        errs.append(f"session block should be one of {VALID_BLOCKS}: {s['block']}")
    return errs


def merge(root: Path, candidates: dict[str, Any], dry_run: bool = False) -> dict[str, Any]:
    comps_path = root / "data" / "competitions.json"
    sched_path = root / "data" / "training-schedule.json"
    report: dict[str, Any] = {"errors": [], "added": {"competitions": 0, "sessions": 0}, "updated": {"competitions": 0, "sessions": 0}}

    incoming_comps = candidates.get("competitions", []) or []
    incoming_sessions = candidates.get("sessions", []) or []

    for c in incoming_comps:
        for e in _validate_competition(c):
            report["errors"].append(e)
    for s in incoming_sessions:
        for e in _validate_session(s):
            report["errors"].append(e)
    if report["errors"]:
        return report

    if incoming_comps:
        current = load_json(comps_path, {"athletes": [], "competitions": []})
        by_id = {c.get("id"): (i, c) for i, c in enumerate(current.get("competitions", []))}
        for c in incoming_comps:
            if c["id"] in by_id:
                idx, _ = by_id[c["id"]]
                current["competitions"][idx] = {**current["competitions"][idx], **c}
                report["updated"]["competitions"] += 1
            else:
                current.setdefault("competitions", []).append(c)
                report["added"]["competitions"] += 1
        if not dry_run:
            dump_json(comps_path, current)

    if incoming_sessions:
        current = load_json(sched_path, {"athletes": [], "sessions": []})
        by_date = {s.get("date"): (i, s) for i, s in enumerate(current.get("sessions", []))}
        for s in incoming_sessions:
            if s["date"] in by_date:
                idx, existing = by_date[s["date"]]
                merged = {**existing}
                for k in ("dow", "course", "block", "focus", "managed_by", "theme"):
                    if s.get(k):
                        merged[k] = s[k]
                current["sessions"][idx] = merged
                report["updated"]["sessions"] += 1
            else:
                current.setdefault("sessions", []).append(s)
                report["added"]["sessions"] += 1
        current["sessions"].sort(key=lambda x: x.get("date", ""))
        if not dry_run:
            dump_json(sched_path, current)

    return report


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--mode", choices=["manual", "file", "url"], help="registration mode")
    p.add_argument("--target", choices=["competitions", "training-schedule", "both"], default="both", help="which target for manual mode (default: both)")
    p.add_argument("--input", help="path (mode=file) or URL (mode=url)")
    p.add_argument("--output", type=Path, help="write candidates JSON to this file (mode=file/url). If omitted → stdout")
    p.add_argument("--merge", type=Path, help="merge a candidates JSON file into the destination JSONs (skip mode)")
    p.add_argument("--dry-run", action="store_true", help="do not write; just report what would change")
    return p


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    root = repo_root()

    if args.merge:
        candidates = load_json(args.merge, {})
        report = merge(root, candidates, dry_run=args.dry_run)
        print(json.dumps(report, ensure_ascii=False, indent=2))
        return 0 if not report["errors"] else 2

    if not args.mode:
        print("error: --mode or --merge required", file=sys.stderr)
        return 1

    if args.mode == "manual":
        candidates = run_manual(root, args.target)
        report = merge(root, candidates, dry_run=args.dry_run)
        print(json.dumps({"candidates": candidates, "merge_report": report}, ensure_ascii=False, indent=2))
        return 0 if not report["errors"] else 2

    if args.mode == "file":
        if not args.input:
            print("error: --input <path> required for mode=file", file=sys.stderr)
            return 1
        raw = run_file(Path(args.input))
    else:  # url
        if not args.input:
            print("error: --input <url> required for mode=url", file=sys.stderr)
            return 1
        raw = run_url(args.input)

    payload = {
        "_note": "Extracted raw content. Review in Workflow F, convert to a candidates JSON with 'competitions' and 'sessions' arrays, then re-run with --merge.",
        "raw": raw,
    }
    text = json.dumps(payload, ensure_ascii=False, indent=2)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(text + "\n", encoding="utf-8")
        print(f"wrote {args.output}")
    else:
        print(text)
    return 0


if __name__ == "__main__":
    sys.exit(main())
